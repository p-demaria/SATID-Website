"""
PORTFOLIO RISK EXPOSURE
=======================
Generates: Portfolio_Risk_Exposure.html

Original module: Tab 2 from consolidated script
Displays: Asset class risk breakdown, portfolio exposure tables
"""

import pandas as pd
import numpy as np
import json
from scipy import stats
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side



# File paths
PORTFOLIO_FILE = 'Model_Portfolio.xlsx'
OHLC_FILE = 'SATID_portfolio_etf_data_weekly_ohlc.csv'
FBIS_PARAMS_FILE = 'SATID_Fbis_Optimized_Parameters.json'
OUTPUT_HTML = 'Portfolio_Risk_Exposure.html'
OUTPUT_EXCEL = 'Portfolio_Risk_Exposure.xlsx'
PORTFOLIO_VALUE = 10_000_000

RISK_THRESHOLDS = {
    'Cash': {'warning': -2.0, 'danger': -3.0},
    'Fixed Income': {'warning': -2.0, 'danger': -3.0},
    'Core Equity': {'warning': -3.5, 'danger': -5.0},
    'Sector & Thematic': {'warning': -3.5, 'danger': -5.0},
    'Secular Growth': {'warning': -3.5, 'danger': -5.0},
    'Alternative Investments': {'warning': -3.5, 'danger': -5.0}
}

COLORS = {'blue': '#34495e', 'orange': '#f39c12', 'red': '#e74c3c'}



def get_active_etfs_with_allocations(portfolio_file):
    df = pd.read_excel(portfolio_file)
    etfs = df[['Unnamed: 2', 'ALLOCATIONS']].copy()
    etfs.columns = ['Ticker', 'Allocation']
    etfs = etfs[etfs['Ticker'].notna()]
    etfs = etfs[etfs['Ticker'] != 'ETF']
    etfs['Allocation'] = pd.to_numeric(etfs['Allocation'], errors='coerce')
    etfs = etfs[etfs['Allocation'] > 0]
    return dict(zip(etfs['Ticker'], etfs['Allocation']))

def calculate_ema(prices, period):
    ema = np.zeros(len(prices))
    ema[0] = prices[0]
    k = 2 / (period + 1)
    for i in range(1, len(prices)):
        ema[i] = prices[i] * k + ema[i-1] * (1 - k)
    return ema



# MODULE 2: Portfolio Risk Exposure Functions
def load_portfolio_allocations(excel_file):
    df = pd.read_excel(excel_file, sheet_name='Portfolio')
    allocations, asset_classes, category_weights = {}, {}, {}
    current_category = None
    
    for idx, row in df.iterrows():
        if idx < 10:
            if pd.notna(row.iloc[1]):
                category_name_raw = str(row.iloc[1]).strip()
                if 'Cash' in category_name_raw and idx == 1:
                    category_weights['Cash'] = float(row['ALLOCATIONS'])
                elif 'Fixed Income' in category_name_raw and idx == 2:
                    category_weights['Fixed Income'] = float(row['ALLOCATIONS'])
                elif 'Core Equity' in category_name_raw and idx == 3:
                    category_weights['Core Equity'] = float(row['ALLOCATIONS'])
                elif 'Sectoral' in category_name_raw and idx == 4:
                    category_weights['Sector & Thematic'] = float(row['ALLOCATIONS'])
                elif 'Secular' in category_name_raw and idx == 5:
                    category_weights['Secular Growth'] = float(row['ALLOCATIONS'])
                elif 'Alternative' in category_name_raw and idx == 6:
                    category_weights['Alternative Investments'] = float(row['ALLOCATIONS'])
            continue
        
        if pd.notna(row.iloc[2]) and str(row.iloc[2]).strip() == 'ETF':
            if pd.notna(row.iloc[0]):
                category_name_raw = str(row.iloc[0]).strip()
                if 'Cash' in category_name_raw:
                    current_category = 'Cash'
                elif 'Fixed Income' in category_name_raw:
                    current_category = 'Fixed Income'
                elif 'Core Equity' in category_name_raw:
                    current_category = 'Core Equity'
                elif 'Sectoral' in category_name_raw or 'Thematic' in category_name_raw:
                    current_category = 'Sector & Thematic'
                elif 'Secular' in category_name_raw:
                    current_category = 'Secular Growth'
                elif 'Alternative' in category_name_raw:
                    current_category = 'Alternative Investments'
            continue
        
        if pd.notna(row.iloc[2]):
            ticker_raw = str(row.iloc[2]).strip()
            if ticker_raw.isupper() and 2 <= len(ticker_raw) <= 5:
                ticker = ticker_raw
                if pd.notna(row['ALLOCATIONS']):
                    try:
                        allocation = float(row['ALLOCATIONS'])
                        if allocation > 0:
                            allocations[ticker] = allocation
                            asset_classes[ticker] = current_category
                    except (ValueError, TypeError):
                        pass
    
    return allocations, asset_classes, category_weights

def calculate_ema_pandas(series, period):
    return series.ewm(span=period, adjust=False).mean()

def calculate_portfolio_risk(df, allocations, asset_classes, params, portfolio_value):
    risk_data = []
    for ticker in allocations.keys():
        close_col = f"{ticker}_close"
        if close_col not in df.columns:
            continue
        prices = df[close_col].dropna()
        if len(prices) < 20:
            continue
        current_price = prices.iloc[-1]
        period = params.get(ticker, {}).get('period', 20)
        shift = params.get(ticker, {}).get('shift', -0.05)
        ema = calculate_ema_pandas(prices, period)
        fbis_support = ema.iloc[-1] * (1 + shift)
        pct_to_support = ((fbis_support - current_price) / current_price) * 100
        allocation_pct = allocations[ticker]
        position_value = portfolio_value * allocation_pct
        usd_to_support = position_value * (pct_to_support / 100)
        asset_class = asset_classes.get(ticker, 'Unknown')
        thresholds = RISK_THRESHOLDS.get(asset_class, {'warning': -3.5, 'danger': -5.0})
        
        if pct_to_support > 0:
            status = 'DANGER'
        elif pct_to_support > thresholds['warning']:
            status = 'WARNING'
        else:
            status = 'SAFE'
        
        risk_data.append({
            'ticker': ticker,
            'asset_class': asset_class,
            'allocation_pct': allocation_pct * 100,
            'position_value': position_value,
            'current_price': current_price,
            'fbis_support': fbis_support,
            'pct_to_support': pct_to_support,
            'usd_to_support': usd_to_support,
            'status': status
        })
    
    risk_data.sort(key=lambda x: x['pct_to_support'])
    
    class_order = ['Cash', 'Fixed Income', 'Core Equity', 'Sector & Thematic', 'Secular Growth', 'Alternative Investments']
    asset_class_summary = {}
    for asset_class in class_order:
        class_positions = [r for r in risk_data if r['asset_class'] == asset_class]
        if class_positions:
            total_allocation = sum(r['allocation_pct'] for r in class_positions)
            total_usd = sum(r['usd_to_support'] for r in class_positions)
            warning_usd = sum(r['usd_to_support'] for r in class_positions if r['status'] == 'WARNING')
            danger_usd = sum(r['usd_to_support'] for r in class_positions if r['status'] == 'DANGER')
            asset_class_summary[asset_class] = {
                'count': len(class_positions),
                'total_allocation': total_allocation,
                'total_usd': total_usd,
                'warning_usd': warning_usd,
                'danger_usd': danger_usd
            }
    
    return risk_data, asset_class_summary

# MODULE 3: SATID Risk Scoring Functions
def calculate_etf_volatility(prices, weeks=13):
    if len(prices) < weeks + 1:
        weeks = len(prices) - 1
    recent_prices = prices[-(weeks + 1):]
    returns = np.diff(recent_prices) / recent_prices[:-1]
    return np.std(returns, ddof=1)

def calculate_satid_score_linear(distance_pct, volatility_weekly, horizon_weeks=1):
    if distance_pct <= 0:
        return 100.0
    volatility_horizon = volatility_weekly * np.sqrt(horizon_weeks)
    z_score = distance_pct / volatility_horizon
    score = 100 - (z_score * 25)
    return max(0, min(100, score))

def analyze_etf_risk(ticker, prices, fbis_params, weeks_lookback=13):
    current_price = prices[-1]
    period = fbis_params.get(ticker, {}).get('period', 8)
    shift = fbis_params.get(ticker, {}).get('shift', 0)
    ema = calculate_ema(prices, period)
    fbis = ema[-1] * (1 + shift)
    distance_pct = (current_price - fbis) / current_price
    volatility = calculate_etf_volatility(prices, weeks_lookback)
    score_1week = calculate_satid_score_linear(distance_pct, volatility, horizon_weeks=1)
    score_1month = calculate_satid_score_linear(distance_pct, volatility, horizon_weeks=4.33)
    
    return {
        'ticker': ticker,
        'current_price': current_price,
        'fbis': fbis,
        'distance_pct': distance_pct,
        'volatility_weekly': volatility,
        'satid_score_1week': score_1week,
        'satid_score_1month': score_1month
    }

def calculate_portfolio_satid_scores(etf_results, allocations):
    portfolio_score_1week = 0
    portfolio_score_1month = 0
    for result in etf_results:
        ticker = result['ticker']
        weight = allocations.get(ticker, 0)
        portfolio_score_1week += result['satid_score_1week'] * weight
        portfolio_score_1month += result['satid_score_1month'] * weight
        result['allocation'] = weight
        result['contribution_1week'] = result['satid_score_1week'] * weight
        result['contribution_1month'] = result['satid_score_1month'] * weight
    return portfolio_score_1week, portfolio_score_1month

def get_risk_level_class(score):
    if score >= 90:
        return 'risk-critical', 'CRITICAL'
    elif score >= 75:
        return 'risk-high', 'HIGH'
    elif score >= 50:
        return 'risk-medium', 'MODERATE'
    elif score >= 25:
        return 'risk-low', 'LOW'
    else:
        return 'risk-minimal', 'MINIMAL'

def get_risk_color_hex(score):
    if score >= 90:
        return 'C0392B'
    elif score >= 75:
        return 'E74C3C'
    elif score >= 50:
        return 'F39C12'
    elif score >= 25:
        return '27AE60'
    else:
        return '16A085'


# ============================================================================
# HTML GENERATION - EXACT ORIGINAL TEMPLATES WRAPPED IN TABS
# ============================================================================

def generate_module2_summary_html(asset_class_summary, category_weights, portfolio_value):
    """Generate Module 2 summary table HTML - EXACT original format"""
    class_order = ['Cash', 'Fixed Income', 'Core Equity', 'Sector & Thematic', 'Secular Growth', 'Alternative Investments']
    portfolio_total_usd = sum(s['total_usd'] for s in asset_class_summary.values())
    portfolio_warning_usd = sum(s['warning_usd'] for s in asset_class_summary.values())
    portfolio_danger_usd = sum(s['danger_usd'] for s in asset_class_summary.values())
    
    html = """
        <table class="summary-table">
            <thead>
                <tr>
                    <th style="width: 30%;">Asset Class</th>
                    <th style="width: 15%; text-align: right;">Allocation</th>
                    <th style="width: 20%; text-align: right;">Amount</th>
                    <th style="width: 35%; text-align: right;">USD to Support</th>
                </tr>
            </thead>
            <tbody>
"""
    
    for asset_class in class_order:
        if asset_class not in asset_class_summary:
            continue
        summary = asset_class_summary[asset_class]
        weight = category_weights.get(asset_class, 0)
        amount = portfolio_value * weight
        html += f"""
                <tr class="class-total-row">
                    <td><strong>{asset_class}</strong></td>
                    <td style="text-align: right;"><strong>{weight*100:.1f}%</strong></td>
                    <td style="text-align: right;"><strong>${amount:,.0f}</strong></td>
                    <td style="text-align: right;"><strong>${summary['total_usd']:,.0f}</strong></td>
                </tr>
                <tr class="summary-warning-row">
                    <td class="label">Warning</td>
                    <td class="value"></td>
                    <td class="value"></td>
                    <td class="warning-value">${summary['warning_usd']:,.0f}</td>
                </tr>
                <tr class="summary-danger-row">
                    <td class="label">Danger</td>
                    <td class="value"></td>
                    <td class="value"></td>
                    <td class="danger-value">${summary['danger_usd']:,.0f}</td>
                </tr>
"""
    
    html += f"""
                <tr class="portfolio-total-row">
                    <td>Portfolio Total</td>
                    <td style="text-align: right;">100.0%</td>
                    <td style="text-align: right;">${portfolio_value:,.0f}</td>
                    <td style="text-align: right;">${portfolio_total_usd:,.0f}</td>
                </tr>
                <tr class="summary-warning-row">
                    <td class="label">Total Warning</td>
                    <td class="value"></td>
                    <td class="value"></td>
                    <td class="warning-value">${portfolio_warning_usd:,.0f}</td>
                </tr>
                <tr class="summary-danger-row">
                    <td class="label">Total Danger</td>
                    <td class="value"></td>
                    <td class="value"></td>
                    <td class="danger-value">${portfolio_danger_usd:,.0f}</td>
                </tr>
            </tbody>
        </table>
"""
    return html

def generate_module2_breakdown_html(risk_data):
    """Generate Module 2 breakdown HTML - EXACT original format"""
    class_order = ['Cash', 'Fixed Income', 'Core Equity', 'Sector & Thematic', 'Secular Growth', 'Alternative Investments']
    html = ""
    
    for asset_class in class_order:
        class_positions = [r for r in risk_data if r['asset_class'] == asset_class]
        if not class_positions:
            continue
        class_positions.sort(key=lambda x: x['pct_to_support'])
        class_warning_usd = sum(r['usd_to_support'] for r in class_positions if r['status'] == 'WARNING')
        class_danger_usd = sum(r['usd_to_support'] for r in class_positions if r['status'] == 'DANGER')
        
        html += f"""
        <div class="breakdown-section">
            <h3>{asset_class}</h3>
            <table class="breakdown-table">
                <thead>
                    <tr>
                        <th>Ticker</th>
                        <th>Allocation</th>
                        <th>Amount</th>
                        <th>Last Price</th>
                        <th>Fbis Support</th>
                        <th>Distance</th>
                        <th>USD to Support</th>
                    </tr>
                </thead>
                <tbody>
"""
        
        for pos in class_positions:
            if pos['status'] == 'WARNING':
                value_html = f'<span class="breakdown-value-warning">${pos["usd_to_support"]:,.0f}</span>'
            elif pos['status'] == 'DANGER':
                value_html = f'<span class="breakdown-value-danger">${pos["usd_to_support"]:,.0f}</span>'
            else:
                value_html = f'<span class="breakdown-value-safe">${pos["usd_to_support"]:,.0f}</span>'
            
            html += f"""
                    <tr>
                        <td class="ticker"><strong>{pos['ticker']}</strong></td>
                        <td>{pos['allocation_pct']:.1f}%</td>
                        <td>${pos['position_value']:,.0f}</td>
                        <td>${pos['current_price']:.2f}</td>
                        <td>${pos['fbis_support']:.2f}</td>
                        <td>{pos['pct_to_support']:.2f}%</td>
                        <td>{value_html}</td>
                    </tr>
"""
        
        html += f"""
                    <tr class="class-summary-warning">
                        <td style="text-align: left; font-weight: bold;">Warning Total</td>
                        <td colspan="5"></td>
                        <td><span class="class-summary-value-warning">${class_warning_usd:,.0f}</span></td>
                    </tr>
                    <tr class="class-summary-danger">
                        <td style="text-align: left; font-weight: bold;">Danger Total</td>
                        <td colspan="5"></td>
                        <td><span class="class-summary-value-danger">${class_danger_usd:,.0f}</span></td>
                    </tr>
                </tbody>
            </table>
        </div>
"""
    return html




def generate_exposure_html(risk_data, asset_class_summary, category_weights, output_file):
    """Generate Portfolio Risk Exposure HTML"""
    
    summary_html = generate_module2_summary_html(asset_class_summary, category_weights, PORTFOLIO_VALUE)
    breakdown_html = generate_module2_breakdown_html(risk_data)
    
    satid_css = """/* SATID Master Stylesheet - Embedded */
* {
    margin: 0;
    padding: 0;
    box-sizing: border-box;
}

body {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    line-height: 1.6;
    color: #2c3e50;
    background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
}

.navbar {
    background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    position: sticky;
    top: 0;
    z-index: 100;
}

.nav-container {
    max-width: 100% !important;
    padding: 0 40px !important;
    display: flex;
    justify-content: space-between;
    align-items: center;
}

.nav-menu {
    display: flex;
    list-style: none;
    gap: 30px;
    justify-content: space-between !important;
    width: 100% !important;
}

.nav-menu li {
    list-style: none;
}

.nav-menu a {
    color: #ecf0f1;
    text-decoration: none;
    padding: 20px 0;
    display: block;
    transition: color 0.3s;
    font-size: 17px;
    font-weight: 400;
}

.nav-menu a:hover,
.nav-menu a.active {
    color: #3498db;
}

.dropdown {
    position: relative;
}

.dropbtn {
    cursor: pointer;
    color: #ecf0f1;
    text-decoration: none;
    padding: 20px 0;
    display: block;
    transition: color 0.3s;
    font-size: 17px;
    font-weight: 400;
}

.dropdown-content {
    display: none;
    position: absolute;
    background-color: #34495e;
    min-width: 250px;
    box-shadow: 0px 8px 16px 0px rgba(0,0,0,0.3);
    z-index: 1000;
    top: 100%;
    left: 0;
}

.dropdown-content a {
    color: #ecf0f1;
    padding: 12px 16px;
    text-decoration: none;
    display: block;
    border-bottom: 1px solid #2c3e50;
}

.dropdown-content a:hover {
    background-color: #2c3e50;
    color: #3498db;
}

.dropdown:hover .dropdown-content {
    display: block;
}

.dropdown:hover .dropbtn {
    color: #3498db;
}

.hero {
    background: linear-gradient(135deg, #1e3c72 0%, #2a5298 50%, #3d6cb9 100%);
    padding: 40px 20px 80px;
    position: relative;
    overflow: hidden;
    color: white;
    text-align: center;
}

.hero::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    bottom: 0;
    background: url('data:image/svg+xml,<svg width="100" height="100" xmlns="http://www.w3.org/2000/svg"><rect width="1" height="1" fill="rgba(255,255,255,0.03)"/></svg>') repeat;
    opacity: 0.5;
}

.hero-content {
    max-width: 1000px;
    margin: 0 auto;
    text-align: center;
    position: relative;
    z-index: 1;
}

.hero h1 {
    font-size: 3.5rem;
    font-weight: 700;
    color: white;
    margin-bottom: 20px;
    letter-spacing: -0.5px;
    animation: fadeInUp 0.8s ease-out;
}

.hero-subtitle {
    font-size: 1.6rem;
    color: rgba(255, 255, 255, 0.9);
    font-weight: 300;
    letter-spacing: 0.5px;
    animation: fadeInUp 0.8s ease-out 0.2s both;
}

@keyframes fadeInUp {
    from {
        opacity: 0;
        transform: translateY(30px);
    }
    to {
        opacity: 1;
        transform: translateY(0);
    }
}

.container {
    max-width: 100% !important;
    margin: -60px auto 60px;
    padding: 0 10px !important;
    position: relative;
    z-index: 2;
}

.wide-content-page {
    background: white;
    border-radius: 16px;
    box-shadow: 0 20px 60px rgba(0, 0, 0, 0.08);
    padding: 50px 30px 70px 30px !important;
    animation: fadeIn 1s ease-out;
    max-width: 1020px;
    margin-left: auto;
    margin-right: auto;
}

@keyframes fadeIn {
    from { opacity: 0; }
    to { opacity: 1; }
}

footer {
    background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
    color: rgba(255, 255, 255, 0.9);
    text-align: center;
    padding: 30px 20px;
    margin-top: 80px;
    font-size: 0.95rem;
    letter-spacing: 0.5px;
}

footer p {
    margin: 0;
}

/* Risk Exposure Specific Styles */
.summary-table {
    width: 100%;
    border-collapse: collapse;
    background-color: white;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    margin-bottom: 30px;
    border-radius: 8px;
    overflow: hidden;
}

.summary-table thead tr {
    background: linear-gradient(135deg, #2a5298 0%, #3d6cb9 100%);
}

.summary-table th {
    color: white;
    padding: 14px 12px;
    font-weight: 600;
    font-size: 1rem;
}

.summary-table td {
    padding: 12px;
    border-bottom: 1px solid #e9ecef;
    font-size: 0.95rem;
}

.class-total-row {
    background: linear-gradient(135deg, #2a5298 0%, #3d6cb9 100%);
    color: white;
    font-weight: 700;
    font-size: 1.05rem;
}

.portfolio-total-row {
    background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
    color: white;
    font-weight: 700;
    font-size: 1.1rem;
}

.summary-warning-row,
.summary-danger-row {
    background: white;
}

.label {
    padding-left: 30px;
    font-style: italic;
    color: #4a5568;
}

.value {
    text-align: right;
}

.warning-value,
.danger-value {
    text-align: right;
    font-weight: 600;
    color: #2c3e50;
}

.breakdown-button {
    display: block;
    width: 320px;
    margin: 30px auto;
    padding: 16px 28px;
    background: linear-gradient(135deg, #2a5298 0%, #3d6cb9 100%);
    color: white;
    border: none;
    border-radius: 8px;
    font-size: 1.05rem;
    font-weight: 600;
    cursor: pointer;
    box-shadow: 0 4px 12px rgba(42, 82, 152, 0.3);
    transition: all 0.3s ease;
    font-family: 'Inter', sans-serif;
}

.breakdown-button:hover {
    transform: translateY(-2px);
    box-shadow: 0 6px 16px rgba(42, 82, 152, 0.4);
    background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
}

.breakdown-details {
    display: none;
    margin-top: 30px;
}

.breakdown-details.visible {
    display: block;
}

.breakdown-section {
    background-color: #f8f9fa;
    padding: 25px;
    margin-bottom: 25px;
    border-radius: 12px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
}

.breakdown-section h3 {
    margin-top: 0;
    margin-bottom: 20px;
    color: #1e3c72;
    font-size: 1.5rem;
    font-weight: 700;
    border-bottom: 3px solid #2a5298;
    padding-bottom: 12px;
}

.breakdown-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.95rem;
    background: white;
    border-radius: 8px;
    overflow: hidden;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04);
}

.breakdown-table thead tr {
    background: linear-gradient(135deg, #2a5298 0%, #3d6cb9 100%);
}

.breakdown-table th {
    color: white;
    padding: 12px 8px;
    font-weight: 600;
    text-align: center;
}

.breakdown-table td {
    padding: 12px 8px;
    border-bottom: 1px solid #e9ecef;
    text-align: center;
}

.breakdown-table tbody tr:hover {
    background-color: #f0f4f8;
}

.ticker {
    font-weight: 700;
    color: #1e3c72;
}

.breakdown-value-warning {
    background: linear-gradient(135deg, #f39c12 0%, #f5a623 100%);
    color: white;
    font-weight: 700;
    padding: 8px 12px;
    text-align: center;
    border-radius: 6px;
    display: inline-block;
    min-width: 90px;
    box-shadow: 0 2px 6px rgba(243, 156, 18, 0.3);
}

.breakdown-value-danger {
    background: linear-gradient(135deg, #e74c3c 0%, #ec7063 100%);
    color: white;
    font-weight: 700;
    padding: 8px 12px;
    text-align: center;
    border-radius: 6px;
    display: inline-block;
    min-width: 90px;
    box-shadow: 0 2px 6px rgba(231, 76, 60, 0.3);
}

.breakdown-value-safe {
    color: #27ae60;
    padding: 8px 12px;
    text-align: center;
    font-weight: 600;
}

@media (max-width: 768px) {
    .hero h1 {
        font-size: 2.5rem;
    }
    
    .hero-subtitle {
        font-size: 1.1rem;
    }
    
    .wide-content-page {
        padding: 30px 20px !important;
    }
    
    .breakdown-button {
        width: 100%;
    }
    
    .nav-menu {
        flex-direction: column;
        gap: 10px;
    }
    
    .dropdown-content {
        position: static;
        box-shadow: none;
        background-color: #2c3e50;
    }
}"""
    
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SATID - Portfolio Risk Exposure</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
{satid_css}
    </style>
</head>
<body>
    <!-- Navigation -->
    <nav class="navbar">
        <div class="nav-container">
            <ul class="nav-menu">
                <li><a href="index.html">Index</a></li>
                <li><a href="Philosophy.html">Philosophy</a></li>
                <li><a href="Methodology.html">Methodology</a></li>
                <li><a href="Support_Levels_Interactive.html">Risk Level Setting</a></li>
                <li><a href="Portfolio_Risk_Exposure.html" class="active">Risk Exposure</a></li>
                <li><a href="SATID_Risk_Score.html">Risk Score</a></li>
                <li><a href="Portfolio_Risk_Dashboard.html">Risk Dashboard</a></li>
                <li class="dropdown">
                    <a href="#" class="dropbtn">Market Views</a>
                    <div class="dropdown-content">
                        <a href="Market_Views.html">About Market Views</a>
                        <a href="SATID_Relative_Performance.html">Cross Asset Validations</a>
                    </div>
                </li>
                <li class="dropdown">
                    <a href="#" class="dropbtn">Appendix</a>
                    <div class="dropdown-content">
                        <a href="Portfolios_Performance.html">Portfolio Performance Analysis</a>
                        <a href="Conventional_Portfolio_Profiles.html">Conventional Portfolio Profiles</a>
                        <a href="model_portfolios.html">Model Portfolios</a>
                        <a href="Portfolio_Statistics.html">Portfolio Statistics</a>
                    </div>
                </li>
            </ul>
        </div>
    </nav>

    <!-- Hero Section -->
    <section class="hero">
        <div class="hero-content">
            <h1>Portfolio Risk Exposure</h1>
            <p class="hero-subtitle">Distance to Fbis support levels by asset class</p>
        </div>
    </section>

    <!-- Main Content Container -->
    <div class="container">
        <div class="wide-content-page">
            {summary_html}
            
            <button class="breakdown-button" onclick="toggleBreakdown()">Show Portfolio Breakdown</button>
            
            <div id="breakdown" class="breakdown-details">
                {breakdown_html}
            </div>
        </div>
    </div>

    <!-- Footer -->
    <footer>
        <p>&copy; 2024 SATID Investment Management. All rights reserved.</p>
    </footer>
    
    <script>
        function toggleBreakdown() {{
            var breakdown = document.getElementById('breakdown');
            var button = document.querySelector('.breakdown-button');
            if (breakdown.classList.contains('visible')) {{
                breakdown.classList.remove('visible');
                button.textContent = 'Show Portfolio Breakdown';
            }} else {{
                breakdown.classList.add('visible');
                button.textContent = 'Hide Portfolio Breakdown';
            }}
        }}
    </script>
</body>
</html>
"""
    
    
    with open(output_file, 'w') as f:
        f.write(html)
    
    print(f"✓ Generated: {output_file}")

def generate_excel_risk_exposure(risk_data, asset_class_summary, category_weights, portfolio_value, output_file):
    """Generate Excel for Portfolio Risk Exposure"""
    print(f"\n📊 Generating Excel: {output_file}...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Exposure"
    
    # Styles
    title_font = Font(size=14, bold=True)
    header_font = Font(size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    class_font = Font(bold=True, size=11)
    class_fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
    warning_fill = PatternFill(start_color='FEF5E7', end_color='FEF5E7', fill_type='solid')
    warning_font = Font(color='F39C12', bold=True)
    warning_border = Border(left=Side(style='thin', color='F39C12'), right=Side(style='thin', color='F39C12'),
                           top=Side(style='thin', color='F39C12'), bottom=Side(style='thin', color='F39C12'))
    danger_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    danger_font = Font(color='E74C3C', bold=True)
    danger_border = Border(left=Side(style='thin', color='E74C3C'), right=Side(style='thin', color='E74C3C'),
                          top=Side(style='thin', color='E74C3C'), bottom=Side(style='thin', color='E74C3C'))
    
    # Title
    ws['A1'] = 'SATID PORTFOLIO RISK EXPOSURE'
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:D2')
    
    # Summary by Asset Class
    row = 4
    ws[f'A{row}'] = 'Summary by Asset Class'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].alignment = center_align
    
    row += 1
    summary_headers = ['Asset Class', 'Allocation', 'Total Value', 'Risk USD']
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    class_order = ['Cash', 'Fixed Income', 'Core Equity', 'Sector & Thematic', 'Secular Growth', 'Alternative Investments']
    
    for asset_class in class_order:
        if asset_class not in asset_class_summary:
            continue
        
        summary = asset_class_summary[asset_class]
        row += 1
        ws[f'A{row}'] = asset_class
        ws[f'B{row}'] = f"{summary['total_allocation']:.1f}%"
        ws[f'C{row}'] = f"${summary['total_usd']:,.0f}"
        ws[f'D{row}'] = f"${summary['warning_usd'] + summary['danger_usd']:,.0f}"
        
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = class_font
            cell.fill = class_fill
            if col > 1:
                cell.alignment = right_align
        
        row += 1
        ws[f'A{row}'] = '  Warning'
        ws[f'D{row}'] = f"${summary['warning_usd']:,.0f}"
        ws[f'D{row}'].alignment = right_align
        
        row += 1
        ws[f'A{row}'] = '  Danger'
        ws[f'D{row}'] = f"${summary['danger_usd']:,.0f}"
        ws[f'D{row}'].alignment = right_align
    
    portfolio_total_usd = sum(s['total_usd'] for s in asset_class_summary.values())
    portfolio_warning_usd = sum(s['warning_usd'] for s in asset_class_summary.values())
    portfolio_danger_usd = sum(s['danger_usd'] for s in asset_class_summary.values())
    
    row += 1
    ws[f'A{row}'] = 'Portfolio Total'
    ws[f'B{row}'] = '100.0%'
    ws[f'C{row}'] = f"${portfolio_value:,.0f}"
    ws[f'D{row}'] = f"${portfolio_total_usd:,.0f}"
    
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.font = class_font
        cell.fill = class_fill
        if col > 1:
            cell.alignment = right_align
    
    row += 1
    ws[f'A{row}'] = '  Total Warning'
    ws[f'D{row}'] = f"${portfolio_warning_usd:,.0f}"
    ws[f'D{row}'].alignment = right_align
    
    row += 1
    ws[f'A{row}'] = '  Total Danger'
    ws[f'D{row}'] = f"${portfolio_danger_usd:,.0f}"
    ws[f'D{row}'].alignment = right_align
    
    # Breakdown by Asset Class
    row += 3
    for asset_class in class_order:
        class_positions = [r for r in risk_data if r['asset_class'] == asset_class]
        if not class_positions:
            continue
        
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = asset_class
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_align
        
        row += 1
        breakdown_headers = ['Ticker', 'Allocation', 'Amount', 'Last Price', 'Fbis Support', 'Distance', 'USD to Support']
        for col, header in enumerate(breakdown_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = class_font
            cell.fill = class_fill
            cell.alignment = center_align
        
        class_positions.sort(key=lambda x: x['pct_to_support'])
        
        for pos in class_positions:
            row += 1
            ws[f'A{row}'] = pos['ticker']
            ws[f'B{row}'] = f"{pos['allocation_pct']:.1f}%"
            ws[f'C{row}'] = f"${pos['position_value']:,.0f}"
            ws[f'D{row}'] = f"${pos['current_price']:.2f}"
            ws[f'E{row}'] = f"${pos['fbis_support']:.2f}"
            ws[f'F{row}'] = f"{pos['pct_to_support']:.2f}%"
            ws[f'G{row}'] = f"${pos['usd_to_support']:,.0f}"
            
            ws[f'A{row}'].font = bold_font
            for col in range(2, 8):
                ws.cell(row=row, column=col).alignment = center_align
            
            if pos['status'] == 'WARNING':
                ws.cell(row=row, column=7).fill = warning_fill
                ws.cell(row=row, column=7).font = warning_font
            elif pos['status'] == 'DANGER':
                ws.cell(row=row, column=7).fill = danger_fill
                ws.cell(row=row, column=7).font = danger_font
        
        class_warning_usd = sum(r['usd_to_support'] for r in class_positions if r['status'] == 'WARNING')
        class_danger_usd = sum(r['usd_to_support'] for r in class_positions if r['status'] == 'DANGER')
        
        row += 1
        ws[f'A{row}'] = 'Warning Total'
        ws[f'G{row}'] = f"${class_warning_usd:,.0f}"
        ws[f'G{row}'].alignment = center_align
        ws[f'G{row}'].fill = warning_fill
        ws[f'G{row}'].font = warning_font
        ws[f'A{row}'].font = bold_font
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = warning_border
        
        row += 1
        ws[f'A{row}'] = 'Danger Total'
        ws[f'G{row}'] = f"${class_danger_usd:,.0f}"
        ws[f'G{row}'].alignment = center_align
        ws[f'G{row}'].fill = danger_fill
        ws[f'G{row}'].font = danger_font
        ws[f'A{row}'].font = bold_font
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = danger_border
        
        row += 2
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 16
    
    wb.save(output_file)
    print(f"  ✓ Excel file saved: {output_file}")


def main():
    print("="*80)
    print("PORTFOLIO RISK EXPOSURE")
    print("="*80)
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    try:
        print("[1/3] Loading portfolio allocations...")
        allocations, asset_classes, category_weights = load_portfolio_allocations(PORTFOLIO_FILE)
        print(f"   ✓ Loaded {len(allocations)} ETFs")
        
        print("[2/3] Loading FBIS parameters...")
        with open(FBIS_PARAMS_FILE, 'r') as f:
            fbis_params = json.load(f)
        print(f"   ✓ Loaded parameters for {len(fbis_params)} ETFs")
        
        print("[3/3] Calculating portfolio risk exposure...")
        df = pd.read_csv(OHLC_FILE, index_col=0, parse_dates=True)
        risk_data, asset_class_summary = calculate_portfolio_risk(
            df, allocations, asset_classes, fbis_params, PORTFOLIO_VALUE
        )
        
        # Generate HTML
        generate_exposure_html(risk_data, asset_class_summary, category_weights, OUTPUT_HTML)
        
        # Generate Excel
        generate_excel_risk_exposure(risk_data, asset_class_summary, category_weights,
                                     PORTFOLIO_VALUE, OUTPUT_EXCEL)
        
        print("\n" + "="*80)
        print("✓ COMPLETE!")
        print("="*80)
        print(f"\nHTML: {OUTPUT_HTML}")
        print(f"Excel: {OUTPUT_EXCEL}")
        print("="*80)
        
    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
