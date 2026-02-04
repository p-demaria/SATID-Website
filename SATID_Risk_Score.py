"""
SATID RISK SCORE DASHBOARD
==========================
Generates: SATID_Risk_Score.html

Original module: Tab 3 from consolidated script
Displays: Individual ETF risk scores, portfolio SATID scores

NOTE: All core calculations imported from SATID_core_calculations.py
      This ensures 100% consistency across all SATID scripts.
"""

import pandas as pd
import numpy as np
import json
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import all core SATID calculation functions from shared module
from SATID_core_calculations import (
    get_active_etfs_with_allocations,
    calculate_ema,
    calculate_portfolio_series,
    calculate_correlation_matrix,
    calculate_individual_volatilities,
    calculate_portfolio_volatility,
    calculate_risk_statistics,
    calculate_etf_volatility,
    calculate_satid_score_linear,
    analyze_etf_risk,
    calculate_portfolio_satid_scores,
    get_risk_level_class
)

# File paths
PORTFOLIO_FILE = 'Model_Portfolio.xlsx'
OHLC_FILE = 'SATID_portfolio_etf_data_weekly_ohlc.csv'
FBIS_PARAMS_FILE = 'SATID_Fbis_Optimized_Parameters.json'
OUTPUT_HTML = 'SATID_Risk_Score.html'
OUTPUT_EXCEL = 'SATID_Risk_Dashboard.xlsx'
PORTFOLIO_VALUE = 10_000_000

RISK_THRESHOLDS = {
    'Cash': {'warning': -2.0, 'danger': -3.0},
    'Fixed Income': {'warning': -2.0, 'danger': -3.0},
    'Core Equity': {'warning': -3.5, 'danger': -5.0},
    'Sector & Thematic': {'warning': -3.5, 'danger': -5.0},
    'Secular Growth': {'warning': -3.5, 'danger': -5.0},
    'Alternative Investments': {'warning': -3.5, 'danger': -5.0}
}

COLORS = {'blue': '#34495e', 'orange': '#f39c12', 'red': '#e74c3c'}


# All core calculation functions imported from SATID_core_calculations.py
# This ensures 100% consistency across all SATID scripts.


def get_risk_color_hex(score):
    """Get hex color code for risk score"""
    if score >= 90:
        return 'C0392B'
    elif score >= 75:
        return 'E74C3C'
    elif score >= 50:
        return 'F39C12'
    elif score >= 25:
        return '27AE60'
    else:
        return '16A085'

def generate_excel_dashboard(etf_results, portfolio_score_1week, portfolio_score_1month, 
                            analysis_date, output_file):
    """Generate Excel dashboard with SATID risk scores"""
    print(f"\n📊 Generating Excel: {output_file}...")
    
    etf_results_sorted = sorted(etf_results, key=lambda x: x['satid_score_1week'], reverse=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "SATID Risk Dashboard"
    
    title_font = Font(size=16, bold=True, color='FFFFFF')
    header_font = Font(size=11, bold=True, color='FFFFFF')
    bold_font = Font(bold=True)
    title_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:J1')
    ws['A1'] = 'SATID RISK SCORING DASHBOARD'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:J2')
    ws['A2'] = f'Support-Adjusted Tactical Investment Discipline | Generated: {analysis_date}'
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = center_align
    
    ws.merge_cells('A4:E4')
    ws['A4'] = 'PORTFOLIO SATID SCORES'
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill
    ws['A4'].alignment = center_align
    
    ws['A5'] = 'Time Horizon'
    ws['B5'] = 'Score'
    ws['C5'] = 'Risk Level'
    ws['D5'] = 'Interpretation'
    ws['E5'] = 'Action Guidance'
    
    for cell in ['A5', 'B5', 'C5', 'D5', 'E5']:
        ws[cell].font = bold_font
        ws[cell].fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
        ws[cell].alignment = center_align
        ws[cell].border = thin_border
    
    _, risk_1w_text = get_risk_level_class(portfolio_score_1week)
    _, risk_1m_text = get_risk_level_class(portfolio_score_1month)
    
    ws['A6'] = '1-Week Horizon'
    ws['B6'] = f"{portfolio_score_1week:.2f}"
    ws['C6'] = risk_1w_text
    
    score_fill_1w = PatternFill(start_color=get_risk_color_hex(portfolio_score_1week), 
                                 end_color=get_risk_color_hex(portfolio_score_1week), 
                                 fill_type='solid')
    ws['B6'].fill = score_fill_1w
    ws['C6'].fill = score_fill_1w
    ws['B6'].font = Font(bold=True, color='FFFFFF')
    ws['C6'].font = Font(bold=True, color='FFFFFF')
    
    ws['A7'] = '1-Month Horizon'
    ws['B7'] = f"{portfolio_score_1month:.2f}"
    ws['C7'] = risk_1m_text
    
    score_fill_1m = PatternFill(start_color=get_risk_color_hex(portfolio_score_1month), 
                                 end_color=get_risk_color_hex(portfolio_score_1month), 
                                 fill_type='solid')
    ws['B7'].fill = score_fill_1m
    ws['C7'].fill = score_fill_1m
    ws['B7'].font = Font(bold=True, color='FFFFFF')
    ws['C7'].font = Font(bold=True, color='FFFFFF')
    
    for row in [6, 7]:
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = thin_border
    
    ws.merge_cells('A9:J9')
    ws['A9'] = 'INDIVIDUAL ETF RISK ANALYSIS'
    ws['A9'].font = header_font
    ws['A9'].fill = header_fill
    ws['A9'].alignment = center_align
    
    headers = ['Ticker', 'Current Price', 'FBIS Support', 'Distance $', 'Distance %', 
               'Weekly Vol', 'Allocation', 'Score (1W)', 'Contrib (1W)', 'Score (1M)', 'Contrib (1M)']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row = 11
    for result in etf_results_sorted:
        ws.cell(row=row, column=1, value=result['ticker']).font = bold_font
        ws.cell(row=row, column=2, value=f"${result['current_price']:.2f}")
        ws.cell(row=row, column=3, value=f"${result['fbis']:.2f}")
        ws.cell(row=row, column=4, value=f"${result['current_price'] - result['fbis']:.2f}")
        ws.cell(row=row, column=5, value=f"{result['distance_pct']:.2%}")
        ws.cell(row=row, column=6, value=f"{result['volatility_weekly']:.2%}")
        ws.cell(row=row, column=7, value=f"{result['allocation']:.2%}")
        ws.cell(row=row, column=8, value=f"{result['satid_score_1week']:.2f}")
        ws.cell(row=row, column=9, value=f"{result['contribution_1week']:.2f}")
        ws.cell(row=row, column=10, value=f"{result['satid_score_1month']:.2f}")
        ws.cell(row=row, column=11, value=f"{result['contribution_1month']:.2f}")
        
        score_1w_fill = PatternFill(start_color=get_risk_color_hex(result['satid_score_1week']), 
                                     end_color=get_risk_color_hex(result['satid_score_1week']), 
                                     fill_type='solid')
        ws.cell(row=row, column=8).fill = score_1w_fill
        ws.cell(row=row, column=8).font = Font(bold=True, color='FFFFFF')
        
        score_1m_fill = PatternFill(start_color=get_risk_color_hex(result['satid_score_1month']), 
                                     end_color=get_risk_color_hex(result['satid_score_1month']), 
                                     fill_type='solid')
        ws.cell(row=row, column=10).fill = score_1m_fill
        ws.cell(row=row, column=10).font = Font(bold=True, color='FFFFFF')
        
        for col in range(1, 12):
            ws.cell(row=row, column=col).alignment = center_align
            ws.cell(row=row, column=col).border = thin_border
        row += 1
    
    ws.column_dimensions['A'].width = 10
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 12
    
    wb.save(output_file)
    print(f"  ✓ Excel file saved: {output_file}")


def generate_risk_score_html(etf_results, portfolio_score_1week, portfolio_score_1month, 
                             allocations, output_file):
    """Generate SATID Risk Score Dashboard HTML"""
    
    # Sort ETFs by risk score
    etf_results_sorted = sorted(etf_results, key=lambda x: x['satid_score_1week'], reverse=True)
    
    # Build table rows
    table_rows = ""
    for result in etf_results_sorted:
        risk_class_1w, _ = get_risk_level_class(result['satid_score_1week'])
        risk_class_1m, _ = get_risk_level_class(result['satid_score_1month'])
        at_support = "⚠️ AT SUPPORT" if result['distance_pct'] <= 0 else ""
        table_rows += f"""
        <tr>
            <td><strong>{result['ticker']}</strong> {at_support}</td>
            <td>${result['current_price']:.2f}</td>
            <td>${result['fbis']:.2f}</td>
            <td>{result['distance_pct']:.2%}</td>
            <td>{result['volatility_weekly']:.2%}</td>
            <td>{result['allocation']:.1%}</td>
            <td class="{risk_class_1w}"><strong>{result['satid_score_1week']:.1f}</strong></td>
            <td>{result['contribution_1week']:.2f}</td>
            <td class="{risk_class_1m}"><strong>{result['satid_score_1month']:.1f}</strong></td>
            <td>{result['contribution_1month']:.2f}</td>
        </tr>
        """
    
    _, risk_label_1w = get_risk_level_class(portfolio_score_1week)
    _, risk_label_1m = get_risk_level_class(portfolio_score_1month)
    
    def get_bg_color(score):
        if score >= 90: return '#c0392b'
        elif score >= 75: return '#e74c3c'
        elif score >= 50: return '#f39c12'
        elif score >= 25: return '#27ae60'
        else: return '#16a085'
    
    bg_color_1w = get_bg_color(portfolio_score_1week)
    bg_color_1m = get_bg_color(portfolio_score_1month)
    analysis_date = datetime.now().strftime("%B %d, %Y at %H:%M")
    
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SATID Risk Score Dashboard</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        /* ============================================
           SATID Website - COMPLETE MASTER STYLESHEET
           ============================================ */

        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            line-height: 1.6;
            color: #2c3e50;
            background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        }}

        /* NAVIGATION */
        .navbar {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            position: sticky;
            top: 0;
            z-index: 100;
        }}

        .nav-container {{
            max-width: 100% !important;
            padding: 0 40px !important;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}

        .nav-menu {{
            display: flex;
            list-style: none;
            gap: 30px;
            justify-content: space-between !important;
            width: 100% !important;
        }}

        .nav-menu li {{
            list-style: none;
        }}

        .nav-menu a {{
            color: #ecf0f1;
            text-decoration: none;
            padding: 20px 0;
            display: block;
            transition: color 0.3s;
            font-size: 17px;
            font-weight: 400;
        }}

        .nav-menu a:hover,
        .nav-menu a.active {{
            color: #3498db;
        }}

        /* DROPDOWN MENU */
        .dropdown {{
            position: relative;
        }}

        .dropbtn {{
            cursor: pointer;
            color: #ecf0f1;
            text-decoration: none;
            padding: 20px 0;
            display: block;
            transition: color 0.3s;
            font-size: 17px;
            font-weight: 400;
        }}

        .dropdown-content {{
            display: none;
            position: absolute;
            background-color: #34495e;
            min-width: 250px;
            box-shadow: 0px 8px 16px 0px rgba(0,0,0,0.3);
            z-index: 1000;
            top: 100%;
            left: 0;
        }}

        .dropdown-content a {{
            color: #ecf0f1;
            padding: 12px 16px;
            text-decoration: none;
            display: block;
            border-bottom: 1px solid #2c3e50;
        }}

        .dropdown-content a:hover {{
            background-color: #2c3e50;
            color: #3498db;
        }}

        .dropdown:hover .dropdown-content {{
            display: block;
        }}

        .dropdown:hover .dropbtn {{
            color: #3498db;
        }}

        /* HERO SECTION */
        .hero {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 50%, #3d6cb9 100%);
            padding: 40px 20px 80px;
            position: relative;
            overflow: hidden;
            color: white;
            text-align: center;
        }}

        .hero::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: url('data:image/svg+xml,<svg width="100" height="100" xmlns="http://www.w3.org/2000/svg"><rect width="1" height="1" fill="rgba(255,255,255,0.03)"/></svg>') repeat;
            opacity: 0.5;
        }}

        .hero-content {{
            max-width: 1000px;
            margin: 0 auto;
            text-align: center;
            position: relative;
            z-index: 1;
        }}

        .hero h1 {{
            font-size: 3.5rem;
            font-weight: 700;
            color: white;
            margin-bottom: 20px;
            letter-spacing: -0.5px;
            animation: fadeInUp 0.8s ease-out;
        }}

        .hero-subtitle {{
            font-size: 1.6rem;
            color: rgba(255, 255, 255, 0.9);
            font-weight: 300;
            letter-spacing: 0.5px;
            animation: fadeInUp 0.8s ease-out 0.2s both;
        }}

        @keyframes fadeInUp {{
            from {{
                opacity: 0;
                transform: translateY(30px);
            }}
            to {{
                opacity: 1;
                transform: translateY(0);
            }}
        }}

        /* CONTAINER */
        .container {{
            max-width: 100% !important;
            margin: -60px auto 60px;
            padding: 0 10px !important;
            position: relative;
            z-index: 2;
        }}

        .content-page {{
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.08);
            padding: 50px 30px 70px 30px !important;
            animation: fadeIn 1s ease-out;
            max-width: 977px;
            margin-left: auto;
            margin-right: auto;
        }}

        @keyframes fadeIn {{
            from {{ opacity: 0; }}
            to {{ opacity: 1; }}
        }}

        /* TYPOGRAPHY */
        h1 {{
            color: #1e3c72;
            font-size: 2.5rem;
            font-weight: 700;
            text-align: center;
            margin-bottom: 10px;
        }}

        h2 {{
            font-size: 2rem;
            font-weight: 600;
            text-align: center;
            margin-bottom: 30px;
            color: #1e3c72;
        }}

        h3 {{
            font-size: 1.4rem;
            font-weight: 600;
            color: #1e3c72;
            margin-top: 35px;
            margin-bottom: 18px;
        }}

        p {{
            font-size: 1.1rem;
            line-height: 1.8;
            color: #4a5568;
            margin-bottom: 20px;
        }}

        /* TABLES */
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
            font-size: 13px;
        }}

        thead {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        }}

        th {{
            padding: 14px 10px;
            text-align: center;
            font-weight: 600;
            font-size: 12px;
            color: white;
        }}

        td {{
            padding: 12px 10px;
            text-align: center;
            border-bottom: 1px solid #ecf0f1;
        }}

        tbody tr:hover {{
            background-color: #e8f0fe;
        }}

        /* RISK LEVEL CLASSES */
        .risk-critical {{ 
            background-color: #c0392b; 
            color: white; 
            font-weight: bold;
            padding: 8px;
            border-radius: 6px;
        }}

        .risk-high {{ 
            background-color: #e74c3c; 
            color: white; 
            font-weight: bold;
            padding: 8px;
            border-radius: 6px;
        }}

        .risk-moderate {{ 
            background-color: #f39c12; 
            color: white; 
            font-weight: bold;
            padding: 8px;
            border-radius: 6px;
        }}

        .risk-low {{ 
            background-color: #27ae60; 
            color: white; 
            font-weight: bold;
            padding: 8px;
            border-radius: 6px;
        }}

        .risk-minimal {{ 
            background-color: #16a085; 
            color: white; 
            font-weight: bold;
            padding: 8px;
            border-radius: 6px;
        }}

        /* SCORE BOXES */
        .scores-container {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 25px;
            margin-bottom: 40px;
        }}

        .score-box {{
            background: white;
            border-radius: 12px;
            padding: 30px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            text-align: center;
        }}

        .score-value {{
            font-size: 30px;
            font-weight: bold;
            color: white;
            margin: 20px 0 15px 0;
            padding: 15px;
            border-radius: 10px;
        }}

        .score-label {{
            font-size: 18px;
            font-weight: bold;
            color: white;
            padding: 12px;
            border-radius: 8px;
            margin-bottom: 10px;
        }}

        .score-description {{
            color: #7f8c8d;
            font-size: 13px;
            margin-top: 10px;
        }}

        .subtitle {{
            text-align: center;
            color: #7f8c8d;
            margin-bottom: 30px;
            font-size: 14px;
        }}

        /* INTERPRETATION SECTIONS */
        .interpretation-section {{
            background: white;
            padding: 25px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}

        .interpretation-grid {{
            display: grid;
            gap: 15px;
        }}

        .interpretation-box {{
            padding: 15px;
            border-left: 4px solid;
            border-radius: 6px;
            background: #f8f9fa;
        }}

        .etf-table-section {{
            background: white;
            padding: 25px;
            border-radius: 10px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            margin-bottom: 20px;
        }}

        /* LEGEND */
        .legend {{
            display: flex;
            justify-content: center;
            gap: 20px;
            margin-top: 20px;
            flex-wrap: wrap;
        }}

        .legend-item {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .legend-box {{
            width: 20px;
            height: 20px;
            border-radius: 4px;
        }}
    </style>
</head>
<body>
    <!-- Navigation -->
    <nav class="navbar">
        <div class="nav-container">
            <ul class="nav-menu">
                <li><a href="index.html">Index</a></li>
                <li><a href="Philosophy.html">Philosophy</a></li>
                <li><a href="Methodology.html">Methodology</a></li>
                <li><a href="Support_Levels_Interactive.html">Risk Level Setting</a></li>
                <li><a href="Portfolio_Risk_Exposure.html">Risk Exposure</a></li>
                <li><a href="SATID_Risk_Score.html" class="active">Risk Score</a></li>
                <li><a href="Portfolio_Risk_Dashboard.html">Risk Dashboard</a></li>
                <li class="dropdown">
                    <a href="#" class="dropbtn">Market Views</a>
                    <div class="dropdown-content">
                        <a href="Market_Views.html">About Market Views</a>
                        <a href="SATID_Relative_Performance.html">Cross Asset Validations</a>
                    </div>
                </li>
                <li class="dropdown">
                    <a href="#" class="dropbtn">Appendix</a>
                    <div class="dropdown-content">
                        <a href="Portfolios_Performance.html">Portfolio Performance Analysis</a>
                        <a href="Conventional_Portfolio_Profiles.html">Conventional Portfolio Profiles</a>
                        <a href="model_portfolios.html">Model Portfolios</a>
                        <a href="Portfolio_Statistics.html">Portfolio Statistics</a>
                    </div>
                </li>
            </ul>
        </div>
    </nav>

    <!-- Hero Section -->
    <section class="hero">
        <div class="hero-content">
            <h1>SATID Risk Score</h1>
            <p class="hero-subtitle">Portfolio Risk Assessment Dashboard</p>
        </div>
    </section>

    <!-- Main Content -->
    <div class="container">
        <div class="content-page">
            
            <div style="text-align: center; color: #7f8c8d; margin-bottom: 30px; font-size: 14px;">
                Analysis: {analysis_date}
            </div>
            
            <div class="scores-container">
            <div class="score-box">
                <h3 style="color: #34495e; margin-bottom: 10px;">Portfolio SATID Score - 1 Week Horizon</h3>
                <div class="score-value" style="background-color: {bg_color_1w};">
                    {portfolio_score_1week:.1f}
                </div>
                <div class="score-label" style="background-color: {bg_color_1w};">
                    {risk_label_1w} RISK
                </div>
                <div class="score-description">
                    Weighted average of individual ETF scores
                </div>
            </div>
            
            <div class="score-box">
                <h3 style="color: #34495e; margin-bottom: 10px;">Portfolio SATID Score - 1 Month Horizon</h3>
                <div class="score-value" style="background-color: {bg_color_1m};">
                    {portfolio_score_1month:.1f}
                </div>
                <div class="score-label" style="background-color: {bg_color_1m};">
                    {risk_label_1m} RISK
                </div>
                <div class="score-description">
                    Adjusted for 1-month volatility (√4.33 weeks)
                </div>
            </div>
        </div>
    
    <div class="interpretation-section">
        <h2>Risk Level Interpretation & Action Guidelines</h2>
        <div class="interpretation-grid">
            <div class="interpretation-box" style="border-left-color: #c0392b; background-color: #fce4e4;">
                <strong style="color: #c0392b;">CRITICAL (90-100)</strong>
                <p style="margin: 5px 0 0 0; font-size: 13px;">0-0.4σ from support. At or touching support - regime change likely. Switch to full SATID mode - actively reduce/exit positions.</p>
            </div>
            <div class="interpretation-box" style="border-left-color: #e74c3c; background-color: #fadbd8;">
                <strong style="color: #e74c3c;">HIGH (75-89)</strong>
                <p style="margin: 5px 0 0 0; font-size: 13px;">0.4-1.0σ from support. Within one standard deviation - one typical move breaches support. High alert - prepare action plan.</p>
            </div>
            <div class="interpretation-box" style="border-left-color: #f39c12; background-color: #fef5e7;">
                <strong style="color: #f39c12;">MODERATE (50-74)</strong>
                <p style="margin: 5px 0 0 0; font-size: 13px;">1.0-2.0σ from support. Need 1-2 standard moves to breach. Monitor closely, consider hybrid MPT/SATID approach.</p>
            </div>
            <div class="interpretation-box" style="border-left-color: #27ae60; background-color: #eafaf1;">
                <strong style="color: #27ae60;">LOW (25-49)</strong>
                <p style="margin: 5px 0 0 0; font-size: 13px;">2.0-3.0σ from support. Comfortable distance from support. Normal MPT management with enhanced monitoring.</p>
            </div>
            <div class="interpretation-box" style="border-left-color: #16a085; background-color: #e8f8f5;">
                <strong style="color: #16a085;">MINIMAL (0-24)</strong>
                <p style="margin: 5px 0 0 0; font-size: 13px;">Beyond 3.0σ from support. Very far from support. Full MPT mode - manage according to benchmark and allocations.</p>
            </div>
        </div>
    </div>
    
    <div class="etf-table-section">
        <h2>Individual ETF Risk Analysis</h2>
        <table>
            <thead>
                <tr>
                    <th>ETF</th>
                    <th>Current Price</th>
                    <th>FBIS Support</th>
                    <th>Distance %</th>
                    <th>Weekly Vol</th>
                    <th>Allocation</th>
                    <th>SATID Score<br>(1 Week)</th>
                    <th>Contribution<br>(1 Week)</th>
                    <th>SATID Score<br>(1 Month)</th>
                    <th>Contribution<br>(1 Month)</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
        
        <div class="legend">
            <div class="legend-item">
                <div class="legend-box" style="background-color: #c0392b;"></div>
                <span>Critical (90-100)</span>
            </div>
            <div class="legend-item">
                <div class="legend-box" style="background-color: #e74c3c;"></div>
                <span>High (75-89)</span>
            </div>
            <div class="legend-item">
                <div class="legend-box" style="background-color: #f39c12;"></div>
                <span>Moderate (50-74)</span>
            </div>
            <div class="legend-item">
                <div class="legend-box" style="background-color: #27ae60;"></div>
                <span>Low (25-49)</span>
            </div>
            <div class="legend-item">
                <div class="legend-box" style="background-color: #16a085;"></div>
                <span>Minimal (0-24)</span>
            </div>
        </div>
    </div>
    
        </div>
    </div>

    <!-- Footer -->
    <footer>
        <p>&copy; 2024 SATID Investment Management. All rights reserved.</p>
    </footer>
</body>
</html>
"""
    
    with open(output_file, 'w') as f:
        f.write(html)
    
    print(f"✓ Generated: {output_file}")

def main():
    print("="*80)
    print("SATID RISK SCORE DASHBOARD")
    print("="*80)
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    try:
        print("[1/4] Loading portfolio allocations...")
        allocations = get_active_etfs_with_allocations(PORTFOLIO_FILE)
        print(f"   ✓ Found {len(allocations)} ETFs")
        
        print("[2/4] Loading FBIS parameters...")
        with open(FBIS_PARAMS_FILE, 'r') as f:
            fbis_params = json.load(f)
        print(f"   ✓ Loaded parameters for {len(fbis_params)} ETFs")
        
        print("[3/4] Analyzing individual ETF risk...")
        df = pd.read_csv(OHLC_FILE)
        df['Date'] = pd.to_datetime(df['Date'])
        df = df.sort_values('Date').reset_index(drop=True)
        
        etf_results = []
        for ticker in allocations.keys():
            close_col = f"{ticker}_close"
            if close_col not in df.columns:
                continue
            prices = df[close_col].values
            result = analyze_etf_risk(ticker, prices, fbis_params)
            etf_results.append(result)
        
        portfolio_score_1week, portfolio_score_1month = calculate_portfolio_satid_scores(
            etf_results, allocations
        )
        
        print("[4/4] Generating HTML dashboard...")
        generate_risk_score_html(etf_results, portfolio_score_1week, portfolio_score_1month,
                                allocations, OUTPUT_HTML)
        
        # Generate Excel
        generate_excel_dashboard(etf_results, portfolio_score_1week, portfolio_score_1month,
                                datetime.now().strftime("%B %d, %Y at %H:%M"), OUTPUT_EXCEL)
        
        print("\n" + "="*80)
        print("✓ COMPLETE!")
        print("="*80)
        print(f"\nHTML: {OUTPUT_HTML}")
        print(f"Excel: {OUTPUT_EXCEL}")
        print(f"\nPortfolio SATID Scores:")
        print(f"  ├─ 1-Week:  {portfolio_score_1week:.2f}")
        print(f"  └─ 1-Month: {portfolio_score_1month:.2f}")
        print("="*80)
        
    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
