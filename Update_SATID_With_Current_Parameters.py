"""
UPDATE SATID WITH CURRENT PARAMETERS
====================================
Master script for WEEKLY routine updates.

This script:
1. LOADS existing parameters from JSON (preserves manual adjustments)
2. Reads updated weekly price data from CSV
3. Regenerates ALL outputs with new prices:
   - Portfolio_Risk_Dashboard.html
   - Portfolio_Risk_Exposure.html + Portfolio_Risk_Exposure.xlsx
   - SATID_Risk_Score.html + SATID_Risk_Dashboard.xlsx
4. Does NOT re-optimize
5. Does NOT overwrite JSON parameters

Use this script: Every week when you update prices and want to refresh all dashboards

Author: SATID Risk Management System
Date: February 2025
"""

import pandas as pd
import numpy as np
import json
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import core calculation functions
from SATID_core_calculations import (
    get_active_etfs_with_allocations,
    calculate_ema,
    calculate_portfolio_series,
    calculate_correlation_matrix,
    calculate_individual_volatilities,
    calculate_portfolio_volatility,
    calculate_risk_statistics,
    calculate_etf_volatility,
    calculate_satid_score_linear,
    analyze_etf_risk,
    calculate_portfolio_satid_scores,
    get_risk_level_class,
    load_portfolio_allocations,
    calculate_portfolio_risk
)

# ================================
# FILE PATHS
# ================================
PORTFOLIO_FILE = 'Model_Portfolio.xlsx'
OHLC_FILE = 'SATID_portfolio_etf_data_weekly_ohlc.csv'
FBIS_PARAMS_FILE = 'SATID_Fbis_Optimized_Parameters.json'
PORTFOLIO_VALUE = 10_000_000

OUTPUT_FILES = {
    'dashboard_html': 'Portfolio_Risk_Dashboard.html',
    'exposure_html': 'Portfolio_Risk_Exposure.html',
    'exposure_xlsx': 'Portfolio_Risk_Exposure.xlsx',
    'risk_score_html': 'SATID_Risk_Score.html',
    'risk_score_xlsx': 'SATID_Risk_Dashboard.xlsx'
}

RISK_THRESHOLDS = {
    'Cash': {'warning': -2.0, 'danger': -3.0},
    'Fixed Income': {'warning': -2.0, 'danger': -3.0},
    'Core Equity': {'warning': -3.5, 'danger': -5.0},
    'Sector & Thematic': {'warning': -3.5, 'danger': -5.0},
    'Secular Growth': {'warning': -3.5, 'danger': -5.0},
    'Alternative Investments': {'warning': -3.5, 'danger': -5.0}
}


# ================================
# MODULE 1: PORTFOLIO RISK DASHBOARD
# ================================
def get_risk_color_and_label(score):
    """Get risk level colors matching SATID methodology"""
    if score >= 90:
        return '#c0392b', "CRITICAL"
    elif score >= 75:
        return '#e74c3c', "HIGH"
    elif score >= 50:
        return '#f39c12', "MODERATE"
    elif score >= 25:
        return '#27ae60', "LOW"
    else:
        return '#16a085', "MINIMAL"


def generate_dashboard_html(dates, portfolio_values, portfolio_fbis,
                           current_value, fbis_value, distance_pct,
                           portfolio_score_1week, portfolio_score_1month,
                           output_file):
    """Generate Portfolio Risk Dashboard HTML"""
    
    chart_dates = [str(d)[:10] for d in dates]
    chart_values = portfolio_values.tolist()
    fbis_dates = [str(d)[:10] for d in dates]
    fbis_values = portfolio_fbis.tolist()
    
    current_nav = current_value * PORTFOLIO_VALUE
    fbis_nav = fbis_value * PORTFOLIO_VALUE
    usd_distance = current_nav - fbis_nav
    distance_pct_display = distance_pct * 100
    
    bg_color_1w, risk_label_1w = get_risk_color_and_label(portfolio_score_1week)
    bg_color_1m, risk_label_1m = get_risk_color_and_label(portfolio_score_1month)
    
    analysis_date = datetime.now().strftime("%B %d, %Y at %H:%M")
    
    html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SATID - Portfolio Risk Dashboard</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            line-height: 1.6;
            color: #2c3e50;
            background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        }}

        .navbar {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            position: sticky;
            top: 0;
            z-index: 1000;
        }}

        .nav-container {{
            max-width: 1800px;
            margin: 0 auto;
            padding: 0 30px;
        }}

        .nav-menu {{
            display: flex;
            flex-wrap: wrap;
            justify-content: center;
            list-style: none;
            padding: 15px 0;
            gap: 5px;
        }}

        .nav-menu li {{
            position: relative;
        }}

        .nav-menu a {{
            color: white;
            text-decoration: none;
            padding: 10px 18px;
            display: block;
            font-weight: 500;
            font-size: 0.95rem;
            transition: all 0.3s ease;
            border-radius: 6px;
        }}

        .nav-menu a:hover {{
            background: rgba(255, 255, 255, 0.15);
        }}

        .nav-menu a.active {{
            background: rgba(255, 255, 255, 0.25);
            font-weight: 600;
        }}

        .dropdown {{
            position: relative;
        }}

        .dropbtn {{
            cursor: pointer;
        }}

        .dropdown-content {{
            display: none;
            position: absolute;
            background-color: #1e3c72;
            min-width: 260px;
            box-shadow: 0px 8px 16px rgba(0,0,0,0.3);
            z-index: 1;
            border-radius: 6px;
            overflow: hidden;
            top: 100%;
            left: 0;
        }}

        .dropdown-content a {{
            padding: 12px 16px;
            display: block;
            border-radius: 0;
        }}

        .dropdown:hover .dropdown-content {{
            display: block;
        }}

        .hero {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            color: white;
            padding: 60px 30px;
            text-align: center;
        }}

        .hero h1 {{
            font-size: 3rem;
            font-weight: 700;
            margin-bottom: 15px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        }}

        .hero-subtitle {{
            font-size: 1.3rem;
            font-weight: 300;
            opacity: 0.95;
        }}

        .container {{
            max-width: 1400px;
            margin: 0 auto;
            padding: 40px 30px;
        }}

        .dashboard-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 25px;
            margin-bottom: 40px;
        }}

        .metric-card {{
            background: white;
            border-radius: 12px;
            padding: 25px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            transition: transform 0.3s ease, box-shadow 0.3s ease;
        }}

        .metric-card:hover {{
            transform: translateY(-5px);
            box-shadow: 0 8px 20px rgba(0,0,0,0.12);
        }}

        .metric-label {{
            font-size: 0.9rem;
            color: #7f8c8d;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 10px;
        }}

        .metric-value {{
            font-size: 2.2rem;
            font-weight: 700;
            color: #2c3e50;
        }}

        .metric-subvalue {{
            font-size: 1.1rem;
            color: #7f8c8d;
            margin-top: 8px;
        }}

        .risk-score-card {{
            background: white;
            border-radius: 12px;
            padding: 25px;
            text-align: center;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            transition: transform 0.3s ease;
        }}

        .risk-score-card:hover {{
            transform: translateY(-5px);
        }}

        .risk-score-value {{
            font-size: 3.5rem;
            font-weight: 800;
            margin: 15px 0;
        }}

        .risk-label {{
            font-size: 1.2rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 1px;
        }}

        .chart-container {{
            background: white;
            border-radius: 12px;
            padding: 30px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            margin-bottom: 30px;
        }}

        .chart-title {{
            font-size: 1.5rem;
            font-weight: 700;
            color: #2c3e50;
            margin-bottom: 20px;
            text-align: center;
        }}

        footer {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            color: white;
            text-align: center;
            padding: 30px;
            margin-top: 60px;
        }}

        footer p {{
            font-size: 0.95rem;
            opacity: 0.9;
        }}

        @media (max-width: 768px) {{
            .hero h1 {{
                font-size: 2rem;
            }}
            
            .hero-subtitle {{
                font-size: 1rem;
            }}
            
            .metric-value {{
                font-size: 1.8rem;
            }}
            
            .risk-score-value {{
                font-size: 2.5rem;
            }}
        }}
    </style>
</head>
<body>
    <!-- Navigation -->
    <nav class="navbar">
        <div class="nav-container">
            <ul class="nav-menu">
                <li><a href="index.html">Index</a></li>
                <li><a href="Philosophy.html">Philosophy</a></li>
                <li><a href="Methodology.html">Methodology</a></li>
                <li><a href="Support_Levels_Interactive.html">Risk Level Setting</a></li>
                <li><a href="Portfolio_Risk_Exposure.html">Risk Exposure</a></li>
                <li><a href="SATID_Risk_Score.html">Risk Score</a></li>
                <li><a href="Portfolio_Risk_Dashboard.html" class="active">Risk Dashboard</a></li>
                <li class="dropdown">
                    <a href="#" class="dropbtn">Market Views</a>
                    <div class="dropdown-content">
                        <a href="Market_Views.html">About Market Views</a>
                        <a href="SATID_Relative_Performance.html">Cross Asset Validations</a>
                    </div>
                </li>
                <li class="dropdown">
                    <a href="#" class="dropbtn">Appendix</a>
                    <div class="dropdown-content">
                        <a href="Portfolios_Performance.html">Portfolio Performance Analysis</a>
                        <a href="Conventional_Portfolio_Profiles.html">Conventional Portfolio Profiles</a>
                        <a href="model_portfolios.html">Model Portfolios</a>
                        <a href="Portfolio_Statistics.html">Portfolio Statistics</a>
                    </div>
                </li>
            </ul>
        </div>
    </nav>

    <!-- Hero Section -->
    <section class="hero">
        <div class="hero-content">
            <h1>Portfolio Risk Dashboard</h1>
            <p class="hero-subtitle">Consolidated portfolio risk analysis and SATID scoring</p>
        </div>
    </section>

    <!-- Main Content -->
    <div class="container">
        <!-- Metrics Grid -->
        <div class="dashboard-grid">
            <div class="metric-card">
                <div class="metric-label">Current Portfolio NAV</div>
                <div class="metric-value">${current_nav:,.0f}</div>
            </div>
            
            <div class="metric-card">
                <div class="metric-label">FBIS Support Level</div>
                <div class="metric-value">${fbis_nav:,.0f}</div>
            </div>
            
            <div class="metric-card">
                <div class="metric-label">Distance to FBIS</div>
                <div class="metric-value">{distance_pct_display:.2f}%</div>
                <div class="metric-subvalue">${usd_distance:,.0f} USD</div>
            </div>
        </div>

        <!-- SATID Scores Grid -->
        <div class="dashboard-grid">
            <div class="risk-score-card" style="border-left: 5px solid {bg_color_1w};">
                <div class="metric-label">SATID Score (1-Week)</div>
                <div class="risk-score-value" style="color: {bg_color_1w};">{portfolio_score_1week:.1f}</div>
                <div class="risk-label" style="color: {bg_color_1w};">{risk_label_1w}</div>
            </div>
            
            <div class="risk-score-card" style="border-left: 5px solid {bg_color_1m};">
                <div class="metric-label">SATID Score (1-Month)</div>
                <div class="risk-score-value" style="color: {bg_color_1m};">{portfolio_score_1month:.1f}</div>
                <div class="risk-label" style="color: {bg_color_1m};">{risk_label_1m}</div>
            </div>
        </div>

        <!-- Portfolio Chart -->
        <div class="chart-container">
            <div class="chart-title">Portfolio Value vs FBIS Support Level</div>
            <div id="portfolioChart"></div>
        </div>

        <!-- Analysis Date -->
        <div style="text-align: center; color: #7f8c8d; font-size: 0.9rem; margin-top: 30px;">
            Analysis Date: {analysis_date}
        </div>
    </div>

    <!-- Footer -->
    <footer>
        <p>&copy; 2024 SATID Investment Management. All rights reserved.</p>
    </footer>

    <script>
        var dates = {chart_dates};
        var portfolioValues = {chart_values};
        var fbisValues = {fbis_values};

        var trace1 = {{
            x: dates,
            y: portfolioValues,
            type: 'scatter',
            mode: 'lines',
            name: 'Portfolio Value',
            line: {{
                color: '#2c3e50',
                width: 3
            }}
        }};

        var trace2 = {{
            x: dates,
            y: fbisValues,
            type: 'scatter',
            mode: 'lines',
            name: 'FBIS Support',
            line: {{
                color: '#e74c3c',
                width: 2,
                dash: 'dot'
            }}
        }};

        var layout = {{
            xaxis: {{
                title: 'Date',
                gridcolor: '#e0e0e0'
            }},
            yaxis: {{
                title: 'Normalized Value',
                gridcolor: '#e0e0e0'
            }},
            hovermode: 'x unified',
            plot_bgcolor: '#f8f9fa',
            paper_bgcolor: 'white',
            font: {{
                family: 'Inter, sans-serif',
                size: 12
            }},
            legend: {{
                orientation: 'h',
                x: 0.5,
                xanchor: 'center',
                y: 1.1
            }},
            height: 500
        }};

        Plotly.newPlot('portfolioChart', [trace1, trace2], layout, {{responsive: true}});
    </script>
</body>
</html>
"""
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"  ✓ Generated: {output_file}")


# ================================
# MODULE 2: PORTFOLIO RISK EXPOSURE
# ================================
def generate_exposure_html(risk_data, asset_class_summary, category_weights, output_file):
    """Generate Portfolio Risk Exposure HTML"""
    
    analysis_date = datetime.now().strftime("%B %d, %Y at %H:%M")
    
    # Sort risk data by % to Support (ascending - most at risk first)
    risk_data_sorted = sorted(risk_data, key=lambda x: x['% to Support'])
    
    # Build asset class summary rows
    asset_class_rows = ""
    for ac in ['Cash', 'Fixed Income', 'Core Equity', 'Sector & Thematic', 'Secular Growth', 'Alternative Investments']:
        if ac in asset_class_summary:
            data = asset_class_summary[ac]
            avg_dist = data['avg_dist']
            
            # Determine color
            thresholds = RISK_THRESHOLDS.get(ac, {'warning': -3.5, 'danger': -5.0})
            if avg_dist <= thresholds['danger']:
                color = '#e74c3c'
            elif avg_dist <= thresholds['warning']:
                color = '#f39c12'
            else:
                color = '#34495e'
            
            asset_class_rows += f"""
            <tr>
                <td style="font-weight: 600;">{ac}</td>
                <td>{data['weight']:.2f}%</td>
                <td style="color: {color}; font-weight: 600;">{avg_dist:.2f}%</td>
                <td>${data['usd_at_risk']:,.0f}</td>
            </tr>
            """
    
    # Build individual ETF rows
    etf_rows = ""
    for item in risk_data_sorted:
        pct = item['% to Support']
        
        # Determine color
        asset_class = item['Asset Class']
        thresholds = RISK_THRESHOLDS.get(asset_class, {'warning': -3.5, 'danger': -5.0})
        if pct <= thresholds['danger']:
            color = '#e74c3c'
        elif pct <= thresholds['warning']:
            color = '#f39c12'
        else:
            color = '#34495e'
        
        etf_rows += f"""
        <tr>
            <td style="font-weight: 600;">{item['Ticker']}</td>
            <td>{item['Asset Class']}</td>
            <td>{item['Allocation (%)']:.2f}%</td>
            <td style="color: {color}; font-weight: 600;">{pct:.2f}%</td>
            <td>${item['USD at Risk']:,.0f}</td>
        </tr>
        """
    
    html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SATID - Portfolio Risk Exposure</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            line-height: 1.6;
            color: #2c3e50;
            background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        }}

        .navbar {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            position: sticky;
            top: 0;
            z-index: 1000;
        }}

        .nav-container {{
            max-width: 1800px;
            margin: 0 auto;
            padding: 0 30px;
        }}

        .nav-menu {{
            display: flex;
            flex-wrap: wrap;
            justify-content: center;
            list-style: none;
            padding: 15px 0;
            gap: 5px;
        }}

        .nav-menu li {{
            position: relative;
        }}

        .nav-menu a {{
            color: white;
            text-decoration: none;
            padding: 10px 18px;
            display: block;
            font-weight: 500;
            font-size: 0.95rem;
            transition: all 0.3s ease;
            border-radius: 6px;
        }}

        .nav-menu a:hover {{
            background: rgba(255, 255, 255, 0.15);
        }}

        .nav-menu a.active {{
            background: rgba(255, 255, 255, 0.25);
            font-weight: 600;
        }}

        .dropdown {{
            position: relative;
        }}

        .dropbtn {{
            cursor: pointer;
        }}

        .dropdown-content {{
            display: none;
            position: absolute;
            background-color: #1e3c72;
            min-width: 260px;
            box-shadow: 0px 8px 16px rgba(0,0,0,0.3);
            z-index: 1;
            border-radius: 6px;
            overflow: hidden;
            top: 100%;
            left: 0;
        }}

        .dropdown-content a {{
            padding: 12px 16px;
            display: block;
            border-radius: 0;
        }}

        .dropdown:hover .dropdown-content {{
            display: block;
        }}

        .hero {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            color: white;
            padding: 60px 30px;
            text-align: center;
        }}

        .hero h1 {{
            font-size: 3rem;
            font-weight: 700;
            margin-bottom: 15px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        }}

        .hero-subtitle {{
            font-size: 1.3rem;
            font-weight: 300;
            opacity: 0.95;
        }}

        .container {{
            max-width: 1400px;
            margin: 0 auto;
            padding: 40px 30px;
        }}

        .section-title {{
            font-size: 1.8rem;
            font-weight: 700;
            color: #2c3e50;
            margin: 40px 0 20px 0;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}

        .table-container {{
            background: white;
            border-radius: 12px;
            padding: 30px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            margin-bottom: 30px;
            overflow-x: auto;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
        }}

        th {{
            background: linear-gradient(135deg, #34495e 0%, #2c3e50 100%);
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: 600;
            font-size: 0.95rem;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #ecf0f1;
        }}

        tr:hover {{
            background-color: #f8f9fa;
        }}

        tr:last-child td {{
            border-bottom: none;
        }}

        footer {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            color: white;
            text-align: center;
            padding: 30px;
            margin-top: 60px;
        }}

        footer p {{
            font-size: 0.95rem;
            opacity: 0.9;
        }}

        .legend {{
            background: #fff3cd;
            border-left: 4px solid #ffc107;
            padding: 15px 20px;
            margin: 20px 0;
            border-radius: 6px;
        }}

        .legend-title {{
            font-weight: 700;
            color: #856404;
            margin-bottom: 10px;
        }}

        .legend-item {{
            color: #856404;
            margin: 5px 0;
            font-size: 0.9rem;
        }}

        @media (max-width: 768px) {{
            .hero h1 {{
                font-size: 2rem;
            }}
            
            .hero-subtitle {{
                font-size: 1rem;
            }}
            
            .section-title {{
                font-size: 1.4rem;
            }}
            
            th, td {{
                padding: 10px;
                font-size: 0.85rem;
            }}
        }}
    </style>
</head>
<body>
    <!-- Navigation -->
    <nav class="navbar">
        <div class="nav-container">
            <ul class="nav-menu">
                <li><a href="index.html">Index</a></li>
                <li><a href="Philosophy.html">Philosophy</a></li>
                <li><a href="Methodology.html">Methodology</a></li>
                <li><a href="Support_Levels_Interactive.html">Risk Level Setting</a></li>
                <li><a href="Portfolio_Risk_Exposure.html" class="active">Risk Exposure</a></li>
                <li><a href="SATID_Risk_Score.html">Risk Score</a></li>
                <li><a href="Portfolio_Risk_Dashboard.html">Risk Dashboard</a></li>
                <li class="dropdown">
                    <a href="#" class="dropbtn">Market Views</a>
                    <div class="dropdown-content">
                        <a href="Market_Views.html">About Market Views</a>
                        <a href="SATID_Relative_Performance.html">Cross Asset Validations</a>
                    </div>
                </li>
                <li class="dropdown">
                    <a href="#" class="dropbtn">Appendix</a>
                    <div class="dropdown-content">
                        <a href="Portfolios_Performance.html">Portfolio Performance Analysis</a>
                        <a href="Conventional_Portfolio_Profiles.html">Conventional Portfolio Profiles</a>
                        <a href="model_portfolios.html">Model Portfolios</a>
                        <a href="Portfolio_Statistics.html">Portfolio Statistics</a>
                    </div>
                </li>
            </ul>
        </div>
    </nav>

    <!-- Hero Section -->
    <section class="hero">
        <div class="hero-content">
            <h1>Portfolio Risk Exposure</h1>
            <p class="hero-subtitle">Asset class risk breakdown and exposure analysis</p>
        </div>
    </section>

    <!-- Main Content -->
    <div class="container">
        <!-- Legend -->
        <div class="legend">
            <div class="legend-title">📊 Risk Thresholds</div>
            <div class="legend-item"><strong style="color: #34495e;">Blue:</strong> Normal (within thresholds)</div>
            <div class="legend-item"><strong style="color: #f39c12;">Orange:</strong> Warning (approaching danger)</div>
            <div class="legend-item"><strong style="color: #e74c3c;">Red:</strong> Danger (exceeded threshold)</div>
        </div>

        <!-- Asset Class Summary -->
        <h2 class="section-title">Asset Class Risk Summary</h2>
        <div class="table-container">
            <table>
                <thead>
                    <tr>
                        <th>Asset Class</th>
                        <th>Allocation</th>
                        <th>Avg. Distance to Support</th>
                        <th>USD at Risk</th>
                    </tr>
                </thead>
                <tbody>
                    {asset_class_rows}
                </tbody>
            </table>
        </div>

        <!-- Individual ETF Exposure -->
        <h2 class="section-title">Individual ETF Exposure</h2>
        <div class="table-container">
            <table>
                <thead>
                    <tr>
                        <th>Ticker</th>
                        <th>Asset Class</th>
                        <th>Allocation</th>
                        <th>% to Support</th>
                        <th>USD at Risk</th>
                    </tr>
                </thead>
                <tbody>
                    {etf_rows}
                </tbody>
            </table>
        </div>

        <!-- Analysis Date -->
        <div style="text-align: center; color: #7f8c8d; font-size: 0.9rem; margin-top: 30px;">
            Analysis Date: {analysis_date}
        </div>
    </div>

    <!-- Footer -->
    <footer>
        <p>&copy; 2024 SATID Investment Management. All rights reserved.</p>
    </footer>
</body>
</html>
"""
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"  ✓ Generated: {output_file}")


def generate_exposure_excel(risk_data, asset_class_summary, category_weights, output_file):
    """Generate Portfolio Risk Exposure Excel"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Exposure"
    
    # Styles
    title_font = Font(size=16, bold=True, color='FFFFFF')
    header_font = Font(size=11, bold=True, color='FFFFFF')
    bold_font = Font(bold=True)
    title_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = 'PORTFOLIO RISK EXPOSURE'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30
    
    # Asset Class Summary
    current_row = 3
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = 'ASSET CLASS SUMMARY'
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].alignment = center_align
    
    current_row += 1
    headers = ['Asset Class', 'Allocation (%)', 'Avg Distance to Support (%)', 'USD at Risk', 'ETF Count']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
        cell.alignment = center_align
        cell.border = thin_border
    
    current_row += 1
    for ac in ['Cash', 'Fixed Income', 'Core Equity', 'Sector & Thematic', 'Secular Growth', 'Alternative Investments']:
        if ac in asset_class_summary:
            data = asset_class_summary[ac]
            ws.cell(row=current_row, column=1, value=ac)
            ws.cell(row=current_row, column=2, value=data['weight']).number_format = '0.00'
            ws.cell(row=current_row, column=3, value=data['avg_dist']).number_format = '0.00'
            ws.cell(row=current_row, column=4, value=data['usd_at_risk']).number_format = '#,##0'
            ws.cell(row=current_row, column=5, value=data['count'])
            
            # Apply color based on risk
            avg_dist = data['avg_dist']
            thresholds = RISK_THRESHOLDS.get(ac, {'warning': -3.5, 'danger': -5.0})
            if avg_dist <= thresholds['danger']:
                fill_color = 'E74C3C'
            elif avg_dist <= thresholds['warning']:
                fill_color = 'F39C12'
            else:
                fill_color = 'FFFFFF'
            
            ws.cell(row=current_row, column=3).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).border = thin_border
            
            current_row += 1
    
    # Individual ETF Exposure
    current_row += 2
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = 'INDIVIDUAL ETF EXPOSURE'
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].alignment = center_align
    
    current_row += 1
    headers = ['Ticker', 'Asset Class', 'Allocation (%)', '% to Support', 'USD at Risk']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
        cell.alignment = center_align
        cell.border = thin_border
    
    current_row += 1
    risk_data_sorted = sorted(risk_data, key=lambda x: x['% to Support'])
    
    for item in risk_data_sorted:
        ws.cell(row=current_row, column=1, value=item['Ticker'])
        ws.cell(row=current_row, column=2, value=item['Asset Class'])
        ws.cell(row=current_row, column=3, value=item['Allocation (%)']).number_format = '0.00'
        ws.cell(row=current_row, column=4, value=item['% to Support']).number_format = '0.00'
        ws.cell(row=current_row, column=5, value=item['USD at Risk']).number_format = '#,##0'
        
        # Apply color based on risk
        pct = item['% to Support']
        asset_class = item['Asset Class']
        thresholds = RISK_THRESHOLDS.get(asset_class, {'warning': -3.5, 'danger': -5.0})
        if pct <= thresholds['danger']:
            fill_color = 'E74C3C'
        elif pct <= thresholds['warning']:
            fill_color = 'F39C12'
        else:
            fill_color = 'FFFFFF'
        
        ws.cell(row=current_row, column=4).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).border = thin_border
        
        current_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    wb.save(output_file)
    print(f"  ✓ Generated: {output_file}")


# ================================
# MODULE 3: SATID RISK SCORE
# ================================
def get_risk_color_hex(score):
    """Get hex color code for risk score"""
    if score >= 90:
        return 'C0392B'
    elif score >= 75:
        return 'E74C3C'
    elif score >= 50:
        return 'F39C12'
    elif score >= 25:
        return '27AE60'
    else:
        return '16A085'


def generate_risk_score_html(etf_results, portfolio_score_1week, portfolio_score_1month, output_file):
    """Generate SATID Risk Score HTML"""
    
    analysis_date = datetime.now().strftime("%B %d, %Y at %H:%M")
    
    # Sort by 1-week score descending
    etf_results_sorted = sorted(etf_results, key=lambda x: x['satid_score_1week'], reverse=True)
    
    # Build ETF rows
    etf_rows = ""
    for result in etf_results_sorted:
        score_1w = result['satid_score_1week']
        score_1m = result['satid_score_1month']
        
        css_class_1w, label_1w = get_risk_level_class(score_1w)
        css_class_1m, label_1m = get_risk_level_class(score_1m)
        
        color_1w = '#' + get_risk_color_hex(score_1w)
        color_1m = '#' + get_risk_color_hex(score_1m)
        
        etf_rows += f"""
        <tr>
            <td style="font-weight: 600;">{result['ticker']}</td>
            <td>{result['allocation']:.2f}%</td>
            <td>${result['current_price']:.2f}</td>
            <td>${result['fbis']:.2f}</td>
            <td>{result['distance_pct']*100:.2f}%</td>
            <td>{result['volatility_weekly']*100:.2f}%</td>
            <td style="background-color: {color_1w}20; color: {color_1w}; font-weight: 600;">{score_1w:.1f}</td>
            <td style="background-color: {color_1m}20; color: {color_1m}; font-weight: 600;">{score_1m:.1f}</td>
        </tr>
        """
    
    # Portfolio scores colors
    portfolio_color_1w = '#' + get_risk_color_hex(portfolio_score_1week)
    portfolio_color_1m = '#' + get_risk_color_hex(portfolio_score_1month)
    portfolio_label_1w = get_risk_level_class(portfolio_score_1week)[1]
    portfolio_label_1m = get_risk_level_class(portfolio_score_1month)[1]
    
    html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SATID - Risk Score Dashboard</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            line-height: 1.6;
            color: #2c3e50;
            background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        }}

        .navbar {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            position: sticky;
            top: 0;
            z-index: 1000;
        }}

        .nav-container {{
            max-width: 1800px;
            margin: 0 auto;
            padding: 0 30px;
        }}

        .nav-menu {{
            display: flex;
            flex-wrap: wrap;
            justify-content: center;
            list-style: none;
            padding: 15px 0;
            gap: 5px;
        }}

        .nav-menu li {{
            position: relative;
        }}

        .nav-menu a {{
            color: white;
            text-decoration: none;
            padding: 10px 18px;
            display: block;
            font-weight: 500;
            font-size: 0.95rem;
            transition: all 0.3s ease;
            border-radius: 6px;
        }}

        .nav-menu a:hover {{
            background: rgba(255, 255, 255, 0.15);
        }}

        .nav-menu a.active {{
            background: rgba(255, 255, 255, 0.25);
            font-weight: 600;
        }}

        .dropdown {{
            position: relative;
        }}

        .dropbtn {{
            cursor: pointer;
        }}

        .dropdown-content {{
            display: none;
            position: absolute;
            background-color: #1e3c72;
            min-width: 260px;
            box-shadow: 0px 8px 16px rgba(0,0,0,0.3);
            z-index: 1;
            border-radius: 6px;
            overflow: hidden;
            top: 100%;
            left: 0;
        }}

        .dropdown-content a {{
            padding: 12px 16px;
            display: block;
            border-radius: 0;
        }}

        .dropdown:hover .dropdown-content {{
            display: block;
        }}

        .hero {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            color: white;
            padding: 60px 30px;
            text-align: center;
        }}

        .hero h1 {{
            font-size: 3rem;
            font-weight: 700;
            margin-bottom: 15px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        }}

        .hero-subtitle {{
            font-size: 1.3rem;
            font-weight: 300;
            opacity: 0.95;
        }}

        .container {{
            max-width: 1600px;
            margin: 0 auto;
            padding: 40px 30px;
        }}

        .portfolio-scores {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 25px;
            margin-bottom: 40px;
        }}

        .score-card {{
            background: white;
            border-radius: 12px;
            padding: 30px;
            text-align: center;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            transition: transform 0.3s ease;
        }}

        .score-card:hover {{
            transform: translateY(-5px);
        }}

        .score-label {{
            font-size: 0.9rem;
            color: #7f8c8d;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 15px;
        }}

        .score-value {{
            font-size: 4rem;
            font-weight: 800;
            margin: 15px 0;
        }}

        .score-risk-label {{
            font-size: 1.2rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 1px;
        }}

        .section-title {{
            font-size: 1.8rem;
            font-weight: 700;
            color: #2c3e50;
            margin: 40px 0 20px 0;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}

        .table-container {{
            background: white;
            border-radius: 12px;
            padding: 30px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            margin-bottom: 30px;
            overflow-x: auto;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
        }}

        th {{
            background: linear-gradient(135deg, #34495e 0%, #2c3e50 100%);
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: 600;
            font-size: 0.9rem;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            white-space: nowrap;
        }}

        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #ecf0f1;
        }}

        tr:hover {{
            background-color: #f8f9fa;
        }}

        tr:last-child td {{
            border-bottom: none;
        }}

        footer {{
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            color: white;
            text-align: center;
            padding: 30px;
            margin-top: 60px;
        }}

        footer p {{
            font-size: 0.95rem;
            opacity: 0.9;
        }}

        @media (max-width: 768px) {{
            .hero h1 {{
                font-size: 2rem;
            }}
            
            .hero-subtitle {{
                font-size: 1rem;
            }}
            
            .score-value {{
                font-size: 2.5rem;
            }}
            
            th, td {{
                padding: 10px;
                font-size: 0.85rem;
            }}
        }}
    </style>
</head>
<body>
    <!-- Navigation -->
    <nav class="navbar">
        <div class="nav-container">
            <ul class="nav-menu">
                <li><a href="index.html">Index</a></li>
                <li><a href="Philosophy.html">Philosophy</a></li>
                <li><a href="Methodology.html">Methodology</a></li>
                <li><a href="Support_Levels_Interactive.html">Risk Level Setting</a></li>
                <li><a href="Portfolio_Risk_Exposure.html">Risk Exposure</a></li>
                <li><a href="SATID_Risk_Score.html" class="active">Risk Score</a></li>
                <li><a href="Portfolio_Risk_Dashboard.html">Risk Dashboard</a></li>
                <li class="dropdown">
                    <a href="#" class="dropbtn">Market Views</a>
                    <div class="dropdown-content">
                        <a href="Market_Views.html">About Market Views</a>
                        <a href="SATID_Relative_Performance.html">Cross Asset Validations</a>
                    </div>
                </li>
                <li class="dropdown">
                    <a href="#" class="dropbtn">Appendix</a>
                    <div class="dropdown-content">
                        <a href="Portfolios_Performance.html">Portfolio Performance Analysis</a>
                        <a href="Conventional_Portfolio_Profiles.html">Conventional Portfolio Profiles</a>
                        <a href="model_portfolios.html">Model Portfolios</a>
                        <a href="Portfolio_Statistics.html">Portfolio Statistics</a>
                    </div>
                </li>
            </ul>
        </div>
    </nav>

    <!-- Hero Section -->
    <section class="hero">
        <div class="hero-content">
            <h1>SATID Risk Score Dashboard</h1>
            <p class="hero-subtitle">Individual ETF risk scores and portfolio SATID analysis</p>
        </div>
    </section>

    <!-- Main Content -->
    <div class="container">
        <!-- Portfolio Scores -->
        <div class="portfolio-scores">
            <div class="score-card" style="border-left: 5px solid {portfolio_color_1w};">
                <div class="score-label">Portfolio SATID Score (1-Week)</div>
                <div class="score-value" style="color: {portfolio_color_1w};">{portfolio_score_1week:.1f}</div>
                <div class="score-risk-label" style="color: {portfolio_color_1w};">{portfolio_label_1w}</div>
            </div>
            
            <div class="score-card" style="border-left: 5px solid {portfolio_color_1m};">
                <div class="score-label">Portfolio SATID Score (1-Month)</div>
                <div class="score-value" style="color: {portfolio_color_1m};">{portfolio_score_1month:.1f}</div>
                <div class="score-risk-label" style="color: {portfolio_color_1m};">{portfolio_label_1m}</div>
            </div>
        </div>

        <!-- Individual ETF Scores -->
        <h2 class="section-title">Individual ETF Risk Scores</h2>
        <div class="table-container">
            <table>
                <thead>
                    <tr>
                        <th>Ticker</th>
                        <th>Allocation</th>
                        <th>Current Price</th>
                        <th>FBIS</th>
                        <th>Distance %</th>
                        <th>Volatility</th>
                        <th>SATID Score<br>(1-Week)</th>
                        <th>SATID Score<br>(1-Month)</th>
                    </tr>
                </thead>
                <tbody>
                    {etf_rows}
                </tbody>
            </table>
        </div>

        <!-- Analysis Date -->
        <div style="text-align: center; color: #7f8c8d; font-size: 0.9rem; margin-top: 30px;">
            Analysis Date: {analysis_date}
        </div>
    </div>

    <!-- Footer -->
    <footer>
        <p>&copy; 2024 SATID Investment Management. All rights reserved.</p>
    </footer>
</body>
</html>
"""
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"  ✓ Generated: {output_file}")


def generate_risk_score_excel(etf_results, portfolio_score_1week, portfolio_score_1month, 
                             analysis_date, output_file):
    """Generate Excel dashboard with SATID risk scores"""
    
    etf_results_sorted = sorted(etf_results, key=lambda x: x['satid_score_1week'], reverse=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "SATID Risk Dashboard"
    
    title_font = Font(size=16, bold=True, color='FFFFFF')
    header_font = Font(size=11, bold=True, color='FFFFFF')
    bold_font = Font(bold=True)
    title_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:J1')
    ws['A1'] = 'SATID RISK SCORING DASHBOARD'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:J2')
    ws['A2'] = f'Analysis Date: {analysis_date}'
    ws['A2'].alignment = center_align
    ws['A2'].font = Font(size=10, italic=True)
    
    # Portfolio Scores
    current_row = 4
    ws.merge_cells(f'A{current_row}:J{current_row}')
    ws[f'A{current_row}'] = 'PORTFOLIO SATID SCORES'
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].alignment = center_align
    
    current_row += 1
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = '1-Week SATID Score'
    ws[f'A{current_row}'].font = bold_font
    ws[f'A{current_row}'].alignment = center_align
    
    ws.merge_cells(f'F{current_row}:J{current_row}')
    ws[f'F{current_row}'] = '1-Month SATID Score'
    ws[f'F{current_row}'].font = bold_font
    ws[f'F{current_row}'].alignment = center_align
    
    current_row += 1
    color_1w = get_risk_color_hex(portfolio_score_1week)
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = round(portfolio_score_1week, 1)
    ws[f'A{current_row}'].font = Font(size=24, bold=True, color=color_1w)
    ws[f'A{current_row}'].fill = PatternFill(start_color=color_1w + '20', end_color=color_1w + '20', fill_type='solid')
    ws[f'A{current_row}'].alignment = center_align
    ws.row_dimensions[current_row].height = 35
    
    color_1m = get_risk_color_hex(portfolio_score_1month)
    ws.merge_cells(f'F{current_row}:J{current_row}')
    ws[f'F{current_row}'] = round(portfolio_score_1month, 1)
    ws[f'F{current_row}'].font = Font(size=24, bold=True, color=color_1m)
    ws[f'F{current_row}'].fill = PatternFill(start_color=color_1m + '20', end_color=color_1m + '20', fill_type='solid')
    ws[f'F{current_row}'].alignment = center_align
    
    # Individual ETF Scores
    current_row += 2
    ws.merge_cells(f'A{current_row}:J{current_row}')
    ws[f'A{current_row}'] = 'INDIVIDUAL ETF SATID SCORES'
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].alignment = center_align
    
    current_row += 1
    headers = ['Ticker', 'Allocation %', 'Current Price', 'FBIS', 'Distance %', 
               'Volatility %', 'SATID Score 1W', 'SATID Score 1M', 
               'Contribution 1W', 'Contribution 1M']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
        cell.alignment = center_align
        cell.border = thin_border
    
    current_row += 1
    for result in etf_results_sorted:
        ws.cell(row=current_row, column=1, value=result['ticker'])
        ws.cell(row=current_row, column=2, value=result['allocation']).number_format = '0.00'
        ws.cell(row=current_row, column=3, value=result['current_price']).number_format = '0.00'
        ws.cell(row=current_row, column=4, value=result['fbis']).number_format = '0.00'
        ws.cell(row=current_row, column=5, value=result['distance_pct']*100).number_format = '0.00'
        ws.cell(row=current_row, column=6, value=result['volatility_weekly']*100).number_format = '0.00'
        ws.cell(row=current_row, column=7, value=result['satid_score_1week']).number_format = '0.0'
        ws.cell(row=current_row, column=8, value=result['satid_score_1month']).number_format = '0.0'
        ws.cell(row=current_row, column=9, value=result['contribution_1week']).number_format = '0.00'
        ws.cell(row=current_row, column=10, value=result['contribution_1month']).number_format = '0.00'
        
        # Color the SATID scores
        score_1w = result['satid_score_1week']
        color_hex_1w = get_risk_color_hex(score_1w)
        ws.cell(row=current_row, column=7).fill = PatternFill(start_color=color_hex_1w + '40', 
                                                               end_color=color_hex_1w + '40', 
                                                               fill_type='solid')
        ws.cell(row=current_row, column=7).font = Font(bold=True, color=color_hex_1w)
        
        score_1m = result['satid_score_1month']
        color_hex_1m = get_risk_color_hex(score_1m)
        ws.cell(row=current_row, column=8).fill = PatternFill(start_color=color_hex_1m + '40', 
                                                               end_color=color_hex_1m + '40', 
                                                               fill_type='solid')
        ws.cell(row=current_row, column=8).font = Font(bold=True, color=color_hex_1m)
        
        for col in range(1, 11):
            ws.cell(row=current_row, column=col).border = thin_border
        
        current_row += 1
    
    # Set column widths
    for col_idx, width in enumerate([12, 12, 13, 13, 12, 12, 15, 15, 15, 15], start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    
    wb.save(output_file)
    print(f"  ✓ Generated: {output_file}")


# ================================
# MAIN EXECUTION
# ================================
def main():
    print("=" * 80)
    print("UPDATE SATID WITH CURRENT PARAMETERS")
    print("=" * 80)
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    print("⚙️  MODE: Weekly Update (using existing parameters)")
    print()
    
    try:
        # Step 1: Load existing parameters from JSON
        print("📂 Loading existing FBIS parameters...")
        with open(FBIS_PARAMS_FILE, 'r') as f:
            fbis_params = json.load(f)
        print(f"  ✓ Loaded parameters for {len(fbis_params)} ETFs from: {FBIS_PARAMS_FILE}")
        print("  ✓ Parameters preserved (no re-optimization)")
        
        # Step 2: Load portfolio data
        print("\n📂 Loading portfolio data...")
        allocations = get_active_etfs_with_allocations(PORTFOLIO_FILE)
        print(f"  ✓ Loaded {len(allocations)} active ETFs from portfolio")
        
        # Load detailed allocations for exposure analysis
        allocations_detailed, asset_classes, category_weights = load_portfolio_allocations(PORTFOLIO_FILE)
        
        # Step 3: Load OHLC data
        print("\n📊 Loading price data...")
        df = pd.read_csv(OHLC_FILE)
        df['Date'] = pd.to_datetime(df['Date'])
        df = df.sort_values('Date').reset_index(drop=True)
        print(f"  ✓ Loaded {len(df)} weeks of price data")
        print(f"  ✓ Latest date: {df['Date'].iloc[-1].strftime('%Y-%m-%d')}")
        
        # Step 4: Calculate portfolio series
        print("\n🔄 Calculating portfolio metrics...")
        dates, portfolio_values, portfolio_fbis = calculate_portfolio_series(
            OHLC_FILE, allocations, fbis_params
        )
        current_value = portfolio_values[-1]
        fbis_value = portfolio_fbis[-1]
        distance_pct = (current_value - fbis_value) / current_value
        
        # Calculate portfolio volatility
        correlation_matrix = calculate_correlation_matrix(OHLC_FILE, list(allocations.keys()), weeks=13)
        individual_vols = calculate_individual_volatilities(OHLC_FILE, list(allocations.keys()), weeks=13)
        portfolio_vol = calculate_portfolio_volatility(allocations, correlation_matrix, individual_vols)
        
        # Calculate SATID scores
        etf_results = []
        for ticker in allocations.keys():
            close_col = f"{ticker}_close"
            if close_col in df.columns:
                prices = df[close_col].values
                result = analyze_etf_risk(ticker, prices, fbis_params, weeks_lookback=13)
                etf_results.append(result)
        
        portfolio_score_1week, portfolio_score_1month = calculate_portfolio_satid_scores(etf_results, allocations)
        
        print(f"  ✓ Portfolio SATID Score (1-Week): {portfolio_score_1week:.1f}")
        print(f"  ✓ Portfolio SATID Score (1-Month): {portfolio_score_1month:.1f}")
        print(f"  ✓ Distance to FBIS: {distance_pct*100:.2f}%")
        
        # Step 5: Generate all outputs
        print("\n📄 Generating HTML dashboards and Excel files...")
        
        # Portfolio Risk Dashboard
        generate_dashboard_html(
            dates, portfolio_values, portfolio_fbis,
            current_value, fbis_value, distance_pct,
            portfolio_score_1week, portfolio_score_1month,
            OUTPUT_FILES['dashboard_html']
        )
        
        # Portfolio Risk Exposure
        risk_data, asset_class_summary = calculate_portfolio_risk(
            df, allocations_detailed, asset_classes, fbis_params, PORTFOLIO_VALUE
        )
        generate_exposure_html(risk_data, asset_class_summary, category_weights, 
                              OUTPUT_FILES['exposure_html'])
        generate_exposure_excel(risk_data, asset_class_summary, category_weights, 
                               OUTPUT_FILES['exposure_xlsx'])
        
        # SATID Risk Score
        analysis_date = datetime.now().strftime("%B %d, %Y at %H:%M")
        generate_risk_score_html(etf_results, portfolio_score_1week, portfolio_score_1month, 
                                OUTPUT_FILES['risk_score_html'])
        generate_risk_score_excel(etf_results, portfolio_score_1week, portfolio_score_1month, 
                                 analysis_date, OUTPUT_FILES['risk_score_xlsx'])
        
        # Summary
        print("\n" + "=" * 80)
        print("✓ WEEKLY UPDATE COMPLETE!")
        print("=" * 80)
        print("\n📄 Generated Files:")
        for key, filename in OUTPUT_FILES.items():
            print(f"  ✓ {filename}")
        
        print("\n🔒 Parameters Status:")
        print(f"  ✓ JSON file preserved: {FBIS_PARAMS_FILE}")
        print(f"  ✓ Manual adjustments retained")
        print(f"  ✓ No re-optimization performed")
        
        print("\n📋 Next steps:")
        print("  1. Review updated dashboards in your browser")
        print("  2. Check Excel files for detailed analysis")
        print("  3. Repeat this script next week when you update prices")
        print("\n💡 To re-optimize parameters, use: Generate_And_Optimize_SATID.py")
        print("=" * 80)
        
    except FileNotFoundError as e:
        print(f"\n✗ ERROR: File not found - {e}")
        print("\nRequired files:")
        print(f"  - {PORTFOLIO_FILE}")
        print(f"  - {OHLC_FILE}")
        print(f"  - {FBIS_PARAMS_FILE}")
        print("\nMake sure all files are in the same directory as this script.")
    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
